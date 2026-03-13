"""
키워드 동향 수집기 v2.0 - GUI 배포용
라이브 채팅 수집 + 대시보드 개선 버전
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import requests
from bs4 import BeautifulSoup
import urllib3
import pandas as pd
from datetime import datetime, timedelta
import time
import random
from collections import Counter
import re
import os
import json
import csv

# 선택적 라이브러리
try:
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    YOUTUBE_AVAILABLE = True
except:
    YOUTUBE_AVAILABLE = False

try:
    from google_play_scraper import reviews_all
    PLAYSTORE_AVAILABLE = True
except:
    PLAYSTORE_AVAILABLE = False

try:
    import pytchat
    PYTCHAT_AVAILABLE = True
except:
    PYTCHAT_AVAILABLE = False

try:
    from google import genai
    GEMINI_AVAILABLE = True
except:
    GEMINI_AVAILABLE = False

try:
    from youtube_transcript_api import YouTubeTranscriptApi
    TRANSCRIPT_AVAILABLE = True
except:
    TRANSCRIPT_AVAILABLE = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ============================================================
# 유튜브 크롤러
# ============================================================
class YouTubeCrawler:
    def __init__(self, api_key):
        self.youtube = build('youtube', 'v3', developerKey=api_key)
    
    def search_videos(self, keyword, max_results=20, callback=None):
        videos_data = []
        try:
            search_response = self.youtube.search().list(
                q=keyword, type='video', part='id,snippet',
                maxResults=max_results, order='viewCount', regionCode='KR'
            ).execute()
            
            video_ids = [item['id']['videoId'] for item in search_response['items']]
            videos_response = self.youtube.videos().list(
                part='snippet,statistics', id=','.join(video_ids)
            ).execute()
            
            channel_ids = [item['snippet']['channelId'] for item in videos_response['items']]
            channels_response = self.youtube.channels().list(
                part='statistics', id=','.join(set(channel_ids))
            ).execute()
            
            channel_subs = {ch['id']: ch['statistics'].get('subscriberCount', '0') 
                          for ch in channels_response['items']}
            
            for idx, item in enumerate(videos_response['items'], 1):
                stats = item['statistics']
                snippet = item['snippet']
                thumbnails = snippet.get('thumbnails', {})
                thumbnail_url = (thumbnails.get('high', {}).get('url') or 
                                thumbnails.get('medium', {}).get('url') or '')
                description = snippet.get('description', '')
                description_short = description[:300] + '...' if len(description) > 300 else description
                
                videos_data.append({
                    '순위': idx, '영상 제목': snippet['title'], '채널명': snippet['channelTitle'],
                    '채널 구독자': self._fmt(channel_subs.get(snippet['channelId'], '0')),
                    '조회수': self._fmt(stats.get('viewCount', '0')),
                    '좋아요': self._fmt(stats.get('likeCount', '0')),
                    '댓글수': self._fmt(stats.get('commentCount', '0')),
                    '업로드일': snippet['publishedAt'][:10],
                    '썸네일': thumbnail_url, '영상 설명': description_short,
                    '링크': f"https://www.youtube.com/watch?v={item['id']}",
                    'video_id': item['id'], '조회수_raw': int(stats.get('viewCount', 0))
                })
                if callback:
                    callback(f"  유튜브 {idx}/{max_results}: {snippet['title'][:30]}...")
            return videos_data
        except Exception as e:
            if callback:
                callback(f"❌ 유튜브 오류: {e}")
            return []
    
    def get_comments(self, video_id, video_title, max_comments=50):
        comments = []
        try:
            response = self.youtube.commentThreads().list(
                part='snippet', videoId=video_id,
                maxResults=min(max_comments, 100),
                order='relevance', textFormat='plainText'
            ).execute()
            for item in response['items']:
                c = item['snippet']['topLevelComment']['snippet']
                comments.append({
                    '영상 제목': video_title, '작성자': c['authorDisplayName'],
                    '댓글': c['textDisplay'], '좋아요': c['likeCount'],
                    '작성일': c['publishedAt'][:10], 'video_id': video_id
                })
            return comments[:max_comments]
        except:
            return []
    
    def get_all_comments(self, video_id, video_title, include_replies=False, callback=None):
        """영상의 모든 댓글 수집 (페이지네이션)"""
        comments = []
        next_page_token = None
        page = 1
        
        try:
            while True:
                request_params = {
                    'part': 'snippet,replies' if include_replies else 'snippet',
                    'videoId': video_id,
                    'maxResults': 100,
                    'order': 'relevance',
                    'textFormat': 'plainText'
                }
                if next_page_token:
                    request_params['pageToken'] = next_page_token
                
                response = self.youtube.commentThreads().list(**request_params).execute()
                
                for item in response['items']:
                    # 최상위 댓글
                    c = item['snippet']['topLevelComment']['snippet']
                    comment_id = item['snippet']['topLevelComment']['id']
                    reply_count = item['snippet'].get('totalReplyCount', 0)
                    
                    comments.append({
                        '영상 제목': video_title,
                        '작성자': c['authorDisplayName'],
                        '댓글': c['textDisplay'],
                        '좋아요': c['likeCount'],
                        '작성일': c['publishedAt'][:10],
                        '유형': '댓글',
                        'video_id': video_id
                    })
                    
                    # 대댓글 수집
                    if include_replies and reply_count > 0:
                        # replies에 일부 대댓글이 포함되어 있음
                        if 'replies' in item:
                            for reply in item['replies']['comments']:
                                r = reply['snippet']
                                comments.append({
                                    '영상 제목': video_title,
                                    '작성자': r['authorDisplayName'],
                                    '댓글': r['textDisplay'],
                                    '좋아요': r['likeCount'],
                                    '작성일': r['publishedAt'][:10],
                                    '유형': '대댓글',
                                    'video_id': video_id
                                })
                        
                        # 대댓글이 5개 이상이면 추가 API 호출 필요
                        if reply_count > 5:
                            additional_replies = self._get_all_replies(comment_id, video_title, video_id)
                            # 이미 가져온 대댓글 제외하고 추가
                            existing_count = len(item.get('replies', {}).get('comments', []))
                            if len(additional_replies) > existing_count:
                                comments.extend(additional_replies[existing_count:])
                
                if callback:
                    callback(f"    댓글 {len(comments)}개 수집 중... (페이지 {page})")
                
                next_page_token = response.get('nextPageToken')
                if not next_page_token:
                    break
                
                page += 1
                time.sleep(0.2)  # API 할당량 보호
                
        except Exception as e:
            if callback:
                callback(f"    ⚠️ 댓글 수집 중 오류: {e}")
        
        return comments
    
    def _get_all_replies(self, parent_id, video_title, video_id):
        """특정 댓글의 모든 대댓글 수집"""
        replies = []
        next_page_token = None
        
        try:
            while True:
                request_params = {
                    'part': 'snippet',
                    'parentId': parent_id,
                    'maxResults': 100,
                    'textFormat': 'plainText'
                }
                if next_page_token:
                    request_params['pageToken'] = next_page_token
                
                response = self.youtube.comments().list(**request_params).execute()
                
                for item in response['items']:
                    r = item['snippet']
                    replies.append({
                        '영상 제목': video_title,
                        '작성자': r['authorDisplayName'],
                        '댓글': r['textDisplay'],
                        '좋아요': r['likeCount'],
                        '작성일': r['publishedAt'][:10],
                        '유형': '대댓글',
                        'video_id': video_id
                    })
                
                next_page_token = response.get('nextPageToken')
                if not next_page_token:
                    break
                
                time.sleep(0.1)
                
        except:
            pass
        
        return replies
    
    def get_video_info_by_id(self, video_id, callback=None):
        """video_id로 영상 정보 가져오기"""
        try:
            response = self.youtube.videos().list(
                part='snippet,statistics',
                id=video_id
            ).execute()
            
            if not response['items']:
                return None
            
            item = response['items'][0]
            stats = item['statistics']
            snippet = item['snippet']
            
            # 채널 구독자 수 가져오기
            channel_response = self.youtube.channels().list(
                part='statistics',
                id=snippet['channelId']
            ).execute()
            
            channel_subs = '0'
            if channel_response['items']:
                channel_subs = channel_response['items'][0]['statistics'].get('subscriberCount', '0')
            
            thumbnails = snippet.get('thumbnails', {})
            thumbnail_url = (thumbnails.get('high', {}).get('url') or 
                            thumbnails.get('medium', {}).get('url') or '')
            description = snippet.get('description', '')
            
            return {
                '영상 제목': snippet['title'],
                '채널명': snippet['channelTitle'],
                '채널 구독자': self._fmt(channel_subs),
                '조회수': self._fmt(stats.get('viewCount', '0')),
                '좋아요': self._fmt(stats.get('likeCount', '0')),
                '댓글수': self._fmt(stats.get('commentCount', '0')),
                '업로드일': snippet['publishedAt'][:10],
                '썸네일': thumbnail_url,
                '영상 설명': description,
                '링크': f"https://www.youtube.com/watch?v={video_id}",
                'video_id': video_id,
                '조회수_raw': int(stats.get('viewCount', 0))
            }
        except Exception as e:
            if callback:
                callback(f"❌ 영상 정보 오류: {e}")
            return None
    
    def _fmt(self, num):
        try:
            return f"{int(num):,}"
        except:
            return str(num)


# ============================================================
# Gemini 영상 요약기 (자막 기반)
# ============================================================
class GeminiVideoSummarizer:
    def __init__(self, api_key):
        if not GEMINI_AVAILABLE:
            raise ImportError("google-genai 라이브러리가 설치되지 않았습니다.\npip install google-genai")
        self.client = genai.Client(api_key=api_key)
        self.model_name = 'gemini-2.5-flash'  # Flash 유지 (Pro는 무료 할당량 거의 없음)
    
    def get_transcript(self, video_id):
        """영상 자막 추출 (자동생성 자막만)"""
        if not TRANSCRIPT_AVAILABLE:
            return None, None, "youtube-transcript-api 미설치"
        
        try:
            ytt_api = YouTubeTranscriptApi()
            transcript_list = ytt_api.list(video_id)
            
            # 자동 생성 자막 (한국어/영어)
            try:
                transcript = transcript_list.find_generated_transcript(['ko', 'en'])
                lang = transcript.language_code
                transcript_data = transcript.fetch()
                full_text = ' '.join([item.text for item in transcript_data.snippets])
                lang_name = "한국어" if 'ko' in lang else "영어"
                return full_text, lang[:2], f"{lang_name} 자동생성 자막"
            except:
                return None, None, "자동생성 자막 없음"
            
        except Exception as e:
            return None, None, f"자막 추출 실패: {str(e)[:30]}"
    
    def summarize_with_transcript(self, transcript_text, language, video_title, callback=None):
        """자막 텍스트 기반 요약 (상세 버전)"""
        try:
            # 자막이 너무 길면 분할 처리
            max_chars = 30000
            if len(transcript_text) > max_chars:
                transcript_text = transcript_text[:max_chars] + "...(이하 생략)"
            
            if language == 'ko':
                prompt = f"""당신은 유튜브 콘텐츠 전문 분석가입니다.

아래는 "{video_title}" 영상의 전체 자막입니다. 이 자막을 바탕으로 영상 내용을 **상세하고 체계적으로** 요약해주세요.

---
[자막 시작]
{transcript_text}
[자막 끝]
---

## 요약 작성 가이드라인:
- 요약만 읽어도 영상을 보지 않고도 내용을 완전히 파악할 수 있어야 합니다
- 영상의 흐름과 구조를 반영해주세요
- 구체적인 수치, 이름, 사례가 있다면 포함해주세요
- 영상 길이에 비례하여 요약 분량을 조절해주세요

## 출력 형식:

### 📺 영상 개요
(이 영상이 무엇에 관한 것인지 2-3문장으로 설명)

### 📝 상세 내용
(영상의 주요 내용을 시간 순서대로 또는 주제별로 정리. 각 항목은 구체적으로 서술)

1. **[첫 번째 주제/섹션]**
   - 세부 내용 설명
   
2. **[두 번째 주제/섹션]**
   - 세부 내용 설명

(필요한 만큼 항목 추가)

### 💡 핵심 포인트
- 영상에서 가장 중요한 메시지나 정보를 bullet point로 정리

### 🎯 결론/요약
(영상의 결론이나 핵심 메시지를 2-3문장으로 정리)

### 🏷️ 키워드
(관련 키워드 5-10개를 쉼표로 구분)"""

            else:
                # 영어 자막인 경우 번역 + 요약
                prompt = f"""당신은 유튜브 콘텐츠 전문 분석가입니다.

아래는 "{video_title}" 영상의 영어 자막입니다. 이 자막을 **한국어로 번역하여** 영상 내용을 **상세하고 체계적으로** 요약해주세요.

---
[영어 자막 시작]
{transcript_text}
[영어 자막 끝]
---

## 요약 작성 가이드라인:
- 요약만 읽어도 영상을 보지 않고도 내용을 완전히 파악할 수 있어야 합니다
- 영상의 흐름과 구조를 반영해주세요
- 구체적인 수치, 이름, 사례가 있다면 포함해주세요
- 영상 길이에 비례하여 요약 분량을 조절해주세요
- 반드시 한국어로 작성해주세요

## 출력 형식:

### 📺 영상 개요
(이 영상이 무엇에 관한 것인지 2-3문장으로 설명)

### 📝 상세 내용
(영상의 주요 내용을 시간 순서대로 또는 주제별로 정리. 각 항목은 구체적으로 서술)

1. **[첫 번째 주제/섹션]**
   - 세부 내용 설명
   
2. **[두 번째 주제/섹션]**
   - 세부 내용 설명

(필요한 만큼 항목 추가)

### 💡 핵심 포인트
- 영상에서 가장 중요한 메시지나 정보를 bullet point로 정리

### 🎯 결론/요약
(영상의 결론이나 핵심 메시지를 2-3문장으로 정리)

### 🏷️ 키워드
(관련 키워드 5-10개를 쉼표로 구분)"""

            response = self.client.models.generate_content(
                model=self.model_name,
                contents=prompt
            )
            
            if response and response.text:
                return response.text, "성공"
            else:
                return None, "요약 생성 실패"
                
        except Exception as e:
            error_msg = str(e)
            return None, f"오류: {error_msg[:100]}"
    
    def analyze_comments_sentiment(self, comments, video_title, callback=None):
        """댓글 감성 분석 (전체 댓글 대상)"""
        try:
            if not comments:
                return None, "댓글 없음"
            
            # 댓글 텍스트 합치기
            comment_texts = []
            for i, c in enumerate(comments):
                comment_text = c.get('댓글', '')
                if comment_text:
                    comment_texts.append(f"{i+1}. {comment_text}")
            
            all_comments_text = "\n".join(comment_texts)
            
            # 토큰 한계 체크 (약 100,000자 = ~25,000 토큰)
            if len(all_comments_text) > 100000:
                return None, f"댓글이 너무 많아 감성 분석 불가 ({len(comments)}개, {len(all_comments_text):,}자)"
            
            prompt = f"""당신은 댓글 데이터 분석 전문가입니다.

아래는 "{video_title}" 영상의 전체 댓글 {len(comments)}개입니다. 
이 댓글들을 정량적으로 분석하고, 주요 주제별로 심층 분석해주세요.

---
[댓글 목록]
{all_comments_text}
---

## 분석 요청사항:

### 📊 정량적 감성 분석

전체 댓글을 긍정/중립/부정으로 분류하고 통계를 제시해주세요.

| 감성 | 개수 | 비율 |
|------|------|------|
| 😊 긍정 | OOO개 | OO% |
| 😐 중립 | OOO개 | OO% |
| 😠 부정 | OOO개 | OO% |
| **합계** | **{len(comments)}개** | **100%** |

---

### 📋 주제별 상세 분석

댓글에서 발견되는 **주요 주제/이슈를 5~10개** 선정하고, 각각에 대해 아래 형식으로 상세 분석해주세요.
(언급 횟수가 많은 순서대로 정렬)

**1. [주제명] (OOO회 언급)**

* **핵심 반응**: 해당 주제에 대한 유저들의 전반적인 반응을 2-3문장으로 요약
* **세부 내용**:
   * 구체적인 댓글 내용이나 패턴 설명
   * 관련된 세부 의견들 정리
   * 이 주제가 시사하는 바

**2. [주제명] (OOO회 언급)**

* **핵심 반응**: ...
* **세부 내용**:
   * ...

(주요 주제 모두 위 형식으로 분석)

---

### 💡 종합 결론

분석 결과를 바탕으로 3-5문장으로 종합 결론을 작성해주세요.
- 시청자/유저들이 가장 중요하게 생각하는 것
- 주요 불만 사항과 개선 요구
- 긍정적으로 평가받는 부분
- 콘텐츠 제작자/운영자에게 도움될 인사이트

---

분석 시 주의사항:
- 정확한 숫자와 비율을 제시해주세요
- 실제 댓글 내용을 인용하여 근거를 제시해주세요
- 객관적이고 균형 잡힌 시각으로 분석해주세요
- 비슷한 의미의 댓글들은 하나의 주제로 묶어주세요"""

            response = self.client.models.generate_content(
                model=self.model_name,
                contents=prompt
            )
            
            if response and response.text:
                return response.text, "성공"
            else:
                return None, "감성 분석 실패"
                
        except Exception as e:
            error_msg = str(e)
            if "429" in error_msg or "quota" in error_msg.lower():
                return None, "API 할당량 초과"
            return None, f"오류: {error_msg[:100]}"
    
    def generate_trend_report(self, keyword, yt_summary, dc_summary, app_summary, play_summary, callback=None):
        """동향 수집 데이터를 종합한 AI 보고서 생성"""
        try:
            # 데이터 요약 텍스트 구성
            data_summary = f"분석 키워드: {keyword}\n분석일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
            
            if yt_summary:
                data_summary += f"=== 📺 유튜브 ===\n"
                data_summary += f"영상 {yt_summary.get('총_영상수', 0)}개, 댓글 {yt_summary.get('총_댓글수', 0)}개 분석\n"
                if yt_summary.get('감성분석'):
                    s = yt_summary['감성분석']
                    data_summary += f"감성: 긍정 {s.get('긍정비율', 'N/A')}, 중립 {s.get('중립비율', 'N/A')}, 부정 {s.get('부정비율', 'N/A')}\n"
                if yt_summary.get('주요_키워드'):
                    kw = ', '.join([f"{k}({c})" for k, c in yt_summary['주요_키워드'][:10]])
                    data_summary += f"주요 키워드: {kw}\n"
                if yt_summary.get('TOP5_영상'):
                    data_summary += "TOP5 영상:\n"
                    for i, v in enumerate(yt_summary['TOP5_영상'][:5], 1):
                        data_summary += f"  {i}. {v.get('영상 제목', '')[:50]} (조회수: {v.get('조회수', 'N/A')})\n"
            
            if dc_summary:
                data_summary += f"\n=== 💬 디시인사이드 ===\n"
                data_summary += f"게시글 {dc_summary.get('총_게시글수', 0)}개 분석\n"
                if dc_summary.get('감성분석'):
                    s = dc_summary['감성분석']
                    data_summary += f"감성: 긍정 {s.get('긍정비율', 'N/A')}, 중립 {s.get('중립비율', 'N/A')}, 부정 {s.get('부정비율', 'N/A')}\n"
                if dc_summary.get('주요_키워드'):
                    kw = ', '.join([f"{k}({c})" for k, c in dc_summary['주요_키워드'][:10]])
                    data_summary += f"주요 키워드: {kw}\n"
            
            if app_summary:
                data_summary += f"\n=== 🍎 앱스토어 ===\n"
                data_summary += f"리뷰 {app_summary.get('총_리뷰수', 0)}개, 평균 평점 {app_summary.get('평균_평점', 0):.1f}점\n"
                if app_summary.get('감성분석'):
                    s = app_summary['감성분석']
                    data_summary += f"감성: 긍정 {s.get('긍정비율', 'N/A')}, 중립 {s.get('중립비율', 'N/A')}, 부정 {s.get('부정비율', 'N/A')}\n"
                if app_summary.get('주요_키워드'):
                    kw = ', '.join([f"{k}({c})" for k, c in app_summary['주요_키워드'][:10]])
                    data_summary += f"주요 키워드: {kw}\n"
            
            if play_summary:
                data_summary += f"\n=== 🤖 플레이스토어 ===\n"
                data_summary += f"리뷰 {play_summary.get('총_리뷰수', 0)}개, 평균 평점 {play_summary.get('평균_평점', 0):.1f}점\n"
                if play_summary.get('감성분석'):
                    s = play_summary['감성분석']
                    data_summary += f"감성: 긍정 {s.get('긍정비율', 'N/A')}, 중립 {s.get('중립비율', 'N/A')}, 부정 {s.get('부정비율', 'N/A')}\n"
                if play_summary.get('주요_키워드'):
                    kw = ', '.join([f"{k}({c})" for k, c in play_summary['주요_키워드'][:10]])
                    data_summary += f"주요 키워드: {kw}\n"
            
            prompt = f'''당신은 시장 동향 분석 전문가입니다.
"{keyword}" 키워드에 대한 멀티플랫폼 동향 분석 보고서를 작성해주세요.

[수집된 데이터 요약]
{data_summary}

## 출력 형식 (반드시 아래 형식을 따라주세요):

# 📊 {keyword} 종합 동향 분석 보고서

## 1. 📈 전체 동향 개요
(전반적인 동향을 3-5문장으로 요약)

## 2. 🎯 플랫폼별 상세 분석

### 📺 유튜브 동향
- 주요 콘텐츠 트렌드
- 시청자 반응 특징
- 주목할 영상/채널

### 💬 커뮤니티 동향
- 주요 논의 주제
- 유저 의견 동향
- 이슈/논란 사항

### 📱 앱스토어 동향
- 전반적인 평가
- 주요 칭찬/불만 사항
- 업데이트 반응

## 3. 😊😐😠 감성 분석 종합
(플랫폼별 감성 비교 및 전체적인 여론 분석)

## 4. 🔑 핵심 키워드 분석
(자주 등장하는 키워드와 그 의미 해석)

## 5. ⚡ 주요 이슈 & 리스크
- 긍정적 이슈
- 부정적 이슈/리스크
- 주의가 필요한 사항

## 6. 💡 시사점 및 제언
(비즈니스/마케팅 관점에서의 인사이트와 제안)

## 7. 📋 핵심 요약 (Executive Summary)
• (핵심 포인트 1)
• (핵심 포인트 2)
• (핵심 포인트 3)
• (핵심 포인트 4)
• (핵심 포인트 5)

---
분석 시 주의사항:
- 데이터에 기반한 객관적 분석을 해주세요
- 구체적인 수치와 비율을 활용해주세요
- 실무에 도움이 되는 액션 아이템을 제시해주세요
- 한국어로 작성해주세요'''

            response = self.client.models.generate_content(
                model=self.model_name,
                contents=prompt
            )
            
            if response and response.text:
                return response.text, "성공"
            else:
                return None, "보고서 생성 실패"
                
        except Exception as e:
            error_msg = str(e)
            if "429" in error_msg or "quota" in error_msg.lower():
                return None, "API 할당량 초과"
            return None, f"오류: {error_msg[:100]}"


# ============================================================
# 라이브 채팅 크롤러
# ============================================================
class YouTubeLiveCrawler:
    def __init__(self):
        if not PYTCHAT_AVAILABLE:
            raise ImportError("pytchat이 설치되지 않았습니다.")
    
    def collect(self, video_id, callback=None, stop_event=None):
        chats = []
        try:
            # interruptable=False로 signal 오류 방지
            chat = pytchat.create(video_id=video_id, interruptable=False)
            if callback:
                callback(f"✅ 유튜브 라이브 연결 성공!")
            
            while chat.is_alive():
                if stop_event and stop_event.is_set():
                    chat.terminate()
                    break
                for item in chat.get().items:
                    chat_data = {
                        '시간': item.datetime, '작성자': item.author.name,
                        '메시지': item.message, '플랫폼': '유튜브 라이브'
                    }
                    chats.append(chat_data)
                    if callback:
                        callback(f"[{item.datetime}] {item.author.name}: {item.message[:30]}...")
                time.sleep(0.5)
        except Exception as e:
            if callback:
                callback(f"❌ 오류: {e}")
        return chats


# ============================================================
# 디시인사이드 크롤러 (IP 차단 리스크 관리 포함)
# ============================================================
class DCInsideCrawler:
    # 브라우저와 유사한 헤더로 봇 차단 완화 (과도한 요청 시 여전히 IP 차단 가능)
    DEFAULT_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://gall.dcinside.com/",
    }
    # 요청 간격(초): 목록 페이지 간격, 글 본문 요청 간격 (차단 완화용)
    DELAY_BETWEEN_PAGES = (1.5, 3.5)
    DELAY_BETWEEN_POSTS = (1.0, 2.5)

    def __init__(self, days_limit=30):
        self.headers = dict(self.DEFAULT_HEADERS)
        self.cutoff_date = datetime.now() - timedelta(days=days_limit)
        self._blocked = False  # 연속 403 감지 시 True → 이후 요청 전부 즉시 중단

    def _parse_date(self, date_str):
        try:
            if '-' in date_str and len(date_str) >= 10:
                return datetime.strptime(date_str[:10], '%Y-%m-%d')
            elif '.' in date_str:
                return datetime.strptime(date_str.split()[0], '%Y.%m.%d')
            return datetime.now()
        except:
            return datetime.now()

    def _request_with_retry(self, url, callback=None):
        """403 시 15초 대기 후 1회 재시도. 재시도도 403이면 IP 차단으로 간주하고 _blocked=True."""
        if self._blocked:
            return None
        try:
            resp = requests.get(url, headers=self.headers, verify=False, timeout=12)
            if resp.status_code == 403:
                if callback:
                    callback("  ⚠️ 디시 접근 제한(403) — 15초 대기 후 재시도")
                time.sleep(random.uniform(15, 20))
                resp = requests.get(url, headers=self.headers, verify=False, timeout=12)
                if resp.status_code == 403:
                    if callback:
                        callback("  ⛔ 디시 IP 차단 확인 (연속 403) — 수집 즉시 중단")
                    self._blocked = True
                    return None
            return resp
        except Exception as e:
            if callback:
                callback(f"  ⚠️ 요청 오류: {e}")
            return None

    def _get_post_content(self, url, callback=None):
        try:
            resp = self._request_with_retry(url, callback)
            if not resp or resp.status_code != 200:
                return ""
            text = resp.text
            if not text or len(text) < 500:
                return ""
            soup = BeautifulSoup(text, "html.parser")
            content_box = soup.select_one("div.writing_view_box")
            return content_box.get_text(separator="\n", strip=True) if content_box else ""
        except Exception:
            return ""
    
    def _get_gallery_posts_from_page(self, soup, gallery_id):
        """갤러리 목록 페이지에서 게시글 행 추출. 여러 선택자 시도."""
        # 1) 기본: 클래스로 게시글 행만
        posts = soup.select("tr.ub-content.us-post")
        if posts:
            return posts
        # 2) tbody tr + td.gall_tit a
        for tr in soup.select("tbody tr"):
            a = tr.select_one("td.gall_tit a[href*='board/view']")
            if not a:
                a = tr.select_one("td.gall_tit a[href*='view']")
            if a:
                href = a.get("href") or ""
                if gallery_id in href and "javascript" not in href:
                    posts.append(tr)
        if posts:
            return posts
        # 3) tbody 없거나 구조 다름: 모든 tr 중 해당 갤 view 링크가 있는 행 (중복 제거)
        seen = set()
        for tr in soup.select("tr"):
            a = tr.select_one("a[href*='view']")
            if not a:
                continue
            href = (a.get("href") or "")
            if gallery_id not in href or "javascript" in href or "id=" not in href:
                continue
            key = id(tr)
            if key in seen:
                continue
            seen.add(key)
            posts.append(tr)
        return posts

    def crawl_gallery(self, gallery_id, is_minor=True, max_pages=10, callback=None):
        results = []
        gallery_type = "mgallery/board" if is_minor else "board"
        for page in range(1, max_pages + 1):
            if self._blocked:
                break
            if page > 1:
                delay = random.uniform(*self.DELAY_BETWEEN_PAGES)
                time.sleep(delay)
            url = f"https://gall.dcinside.com/{gallery_type}/lists/?id={gallery_id}&page={page}"
            try:
                resp = self._request_with_retry(url, callback)
                if not resp or resp.status_code != 200:
                    if callback:
                        callback(f"  디시 갤러리 {page}페이지: 응답 실패 (IP 제한 가능성)")
                    continue
                html = resp.text
                soup = BeautifulSoup(html, "html.parser")
                posts = self._get_gallery_posts_from_page(soup, gallery_id)
                if not posts:
                    if callback:
                        if len(html) < 2000 or gallery_id not in html:
                            callback(f"  디시 갤러리 {page}페이지: 목록 0개 (IP 제한·캡차 또는 비정상 응답 가능성)")
                        else:
                            callback(f"  디시 갤러리 {page}페이지: 목록 0개 (HTML 구조 변경 가능성)")
                time.sleep(random.uniform(1.0, 2.0))
                added = 0
                for post in posts:
                    title_el = post.select_one("td.gall_tit a")
                    if not title_el:
                        for a in post.select("a[href*='view']"):
                            h = a.get("href") or ""
                            if gallery_id in h and "javascript" not in h and "id=" in h:
                                title_el = a
                                break
                    date_el = post.select_one("td.gall_date")
                    if not title_el:
                        continue
                    href = title_el.get("href") or ""
                    if "javascript" in href or gallery_id not in href:
                        continue
                    date_str = date_el.get("title", date_el.text.strip()) if date_el else ""
                    link = "https://gall.dcinside.com" + href if href.startswith("/") else href
                    if not link.startswith("http"):
                        link = "https://gall.dcinside.com" + link
                    content = self._get_post_content(link, callback)
                    results.append({
                        "제목": title_el.text.strip(), "링크": link,
                        "본문": content, "작성일": date_str, "출처": f"갤러리({gallery_id})"
                    })
                    added += 1
                    time.sleep(random.uniform(*self.DELAY_BETWEEN_POSTS))
                if callback:
                    callback(f"  디시 갤러리 {page}페이지: {added}개 수집 (총 {len(results)}개)")
            except Exception as e:
                if callback:
                    callback(f"  디시 갤러리 {page}페이지 요청 오류: {e}")
                break
        return results
    
    def crawl_search(self, keyword, max_pages=10, callback=None):
        results = []
        for page in range(1, max_pages + 1):
            if self._blocked:
                break
            if page > 1:
                time.sleep(random.uniform(*self.DELAY_BETWEEN_PAGES))
            url = f"https://search.dcinside.com/post/p/{page}/q/{keyword}"
            try:
                resp = self._request_with_retry(url, callback)
                if not resp or resp.status_code != 200:
                    if callback:
                        callback(f"  디시 검색 {page}페이지: 응답 실패 (IP 제한 가능성)")
                    continue
                soup = BeautifulSoup(resp.text, "html.parser")
                posts = soup.select("ul.sch_result_list > li")
                if not posts:
                    break
                time.sleep(random.uniform(1.0, 2.0))
                for post in posts:
                    title_el = post.select_one("a.tit_txt")
                    date_el = post.select_one("span.date")
                    if title_el:
                        date_str = date_el.text.strip() if date_el else ""
                        post_date = self._parse_date(date_str)
                        if post_date < self.cutoff_date:
                            continue
                        link = title_el['href']
                        content = self._get_post_content(link, callback)
                        results.append({
                            '제목': title_el.text.strip(), '링크': link,
                            '본문': content, '작성일': date_str, '출처': '키워드검색'
                        })
                        time.sleep(random.uniform(*self.DELAY_BETWEEN_POSTS))
                if callback:
                    callback(f"  디시 검색 {page}페이지 완료")
            except Exception:
                break
        return results
    
    def crawl_all(self, keyword, gallery_id=None, is_minor=True, max_pages=10,
                  do_keyword=True, do_gallery=True, callback=None):
        """
        디시인사이드 수집: 키워드 검색과 갤러리 수집을 선택적으로 수행.
        - do_keyword: True면 키워드 검색 결과 수집 (crawl_search)
        - do_gallery: True면 특정 갤러리 글 수집 (crawl_gallery, gallery_id 필요)
        """
        all_posts = []
        if do_gallery and gallery_id:
            all_posts.extend(self.crawl_gallery(gallery_id, is_minor, max_pages, callback))
        if not self._blocked and do_keyword and keyword:
            all_posts.extend(self.crawl_search(keyword, max_pages, callback))
        seen = set()
        unique = [p for p in all_posts if p['링크'] not in seen and not seen.add(p['링크'])]
        return unique


# ============================================================
# 앱스토어/플레이스토어 크롤러
# ============================================================
class AppStoreCrawler:
    def crawl(self, app_id, max_pages=10, callback=None):
        all_reviews = []
        for sort_by in ['mostrecent', 'mostHelpful']:
            for page in range(1, max_pages + 1):
                url = f"https://itunes.apple.com/kr/rss/customerreviews/page={page}/id={app_id}/sortby={sort_by}/xml"
                try:
                    resp = requests.get(url, verify=False, timeout=15)
                    soup = BeautifulSoup(resp.content, 'lxml-xml')
                    entries = soup.find_all('entry')
                    if not entries:
                        break
                    for entry in entries:
                        all_reviews.append({
                            'review_id': entry.find('id').text if entry.find('id') else '',
                            '제목': entry.find('title').text if entry.find('title') else '',
                            '작성자': entry.find('author').find('name').text if entry.find('author') else '',
                            '평점': entry.find('im:rating').text if entry.find('im:rating') else '',
                            '내용': entry.find('content').text if entry.find('content') else '',
                            '작성일': entry.find('updated').text[:10] if entry.find('updated') else '',
                            '버전': entry.find('im:version').text if entry.find('im:version') else '',
                        })
                    time.sleep(0.5)
                except:
                    break
            if callback:
                callback(f"  앱스토어 {sort_by} 수집 완료")
        df = pd.DataFrame(all_reviews)
        if 'review_id' in df.columns and len(df) > 0:
            df = df.drop_duplicates(subset=['review_id'])
        return df.to_dict('records') if len(df) > 0 else []


class PlayStoreCrawler:
    def crawl(self, app_id, max_reviews=500, callback=None):
        try:
            if callback:
                callback("  플레이스토어 수집 중... (시간이 걸릴 수 있습니다)")
            result = reviews_all(app_id, lang='ko', country='kr')
            reviews = [{'작성자': r.get('userName', ''), '평점': r.get('score', ''),
                       '내용': r.get('content', ''), '작성일': str(r.get('at', ''))[:10],
                       '좋아요': r.get('thumbsUpCount', 0)} for r in result[:max_reviews]]
            return reviews
        except Exception as e:
            if callback:
                callback(f"❌ 플레이스토어 오류: {e}")
            return []


# ============================================================
# 데이터 분석기 (개선)
# ============================================================
class DataAnalyzer:
    POSITIVE_KEYWORDS = [
        '좋아', '좋은', '좋다', '좋음', '좋네', '재밌', '재미', '재밌다', '최고', '최고다',
        '굿', 'good', 'nice', 'great', '추천', '강추', '적극추천', '만족', '대만족',
        '잘만든', '꿀잼', '핵꿀잼', '존잼', '갓겜', '명작', '수작', '최애', '사랑',
        '꿀팁', '유용', '혜자', '이벤트', '보상', '안정', '퀄리티', '중독', '대박',
        '감동', '감사', '기대', '설렘', '완벽', '훌륭', '멋진', '예쁜', '신규', '업데이트'
    ]
    
    NEGATIVE_KEYWORDS = [
        '별로', '별로다', '노잼', '핵노잼', '쓰레기', '쓰렉', '최악', '짜증', '빡침',
        '버그', '오류', '튕김', '렉', '망겜', '망작', '폭망', '환불', '삭제', '탈퇴',
        '거지', '노답', '답없', '현질', '과금', '페이투윈', '불만', '실망', '후회',
        '하향', '너프', '핵', '어뷰징', '확률', '뽑기', '가챠', '광고', '느림', '로딩',
        '극혐', '혐오', '싫어', '안해', '못해', '지움', '끊김', '멈춤', '심각', '문제'
    ]
    
    # 주제 그룹핑 (동향 맥락용)
    TOPIC_GROUPS = {
        '기술적 이슈': ['버그', '오류', '렉', '튕김', '끊김', '로딩', '느림', '멈춤', '크래시'],
        '과금/경제': ['과금', '현질', '확률', '뽑기', '가챠', '페이투윈', '비싸', '돈', '결제'],
        '콘텐츠/재미': ['재미', '재밌', '콘텐츠', '스토리', '캐릭터', '이벤트', '업데이트', '신규'],
        '밸런스': ['밸런스', '너프', '하향', '상향', '버프', '밸붕', '불균형'],
        '고객응대': ['고객', '응대', 'CS', '답변', '환불', '문의', '서비스'],
        '그래픽/UI': ['그래픽', '연출', 'UI', '인터페이스', '디자인', '예쁜', '퀄리티']
    }
    
    @staticmethod
    def extract_keywords(texts, top_n=20):
        stopwords = {'있는', '하는', '있고', '있어', '하고', '에서', '으로', '이런', '저런',
                    '그런', '이게', '저게', '그게', '진짜', '너무', '정말', '아니', '이거',
                    '하면', '있음', '없음', '같은', '된다', '한다', '그리고', '하지만'}
        words = []
        for text in texts:
            if not text:
                continue
            korean = re.findall(r'[가-힣]+', str(text))
            words.extend([w for w in korean if len(w) >= 2 and w not in stopwords])
        return Counter(words).most_common(top_n)
    
    @staticmethod
    def analyze_sentiment(texts):
        pos_count = neg_count = neu_count = 0
        pos_found, neg_found = [], []
        
        for text in texts:
            text_lower = str(text).lower()
            is_pos = any(p in text_lower for p in DataAnalyzer.POSITIVE_KEYWORDS)
            is_neg = any(n in text_lower for n in DataAnalyzer.NEGATIVE_KEYWORDS)
            
            for p in DataAnalyzer.POSITIVE_KEYWORDS:
                if p in text_lower:
                    pos_found.append(p)
            for n in DataAnalyzer.NEGATIVE_KEYWORDS:
                if n in text_lower:
                    neg_found.append(n)
            
            if is_pos and not is_neg:
                pos_count += 1
            elif is_neg and not is_pos:
                neg_count += 1
            else:
                neu_count += 1
        
        total = len(texts) if texts else 1
        return {
            '긍정': pos_count, '부정': neg_count, '중립': neu_count,
            '긍정비율': f"{pos_count/total*100:.1f}%",
            '부정비율': f"{neg_count/total*100:.1f}%",
            '중립비율': f"{neu_count/total*100:.1f}%",
            '긍정비율_num': pos_count/total*100, '부정비율_num': neg_count/total*100,
            '긍정키워드': Counter(pos_found).most_common(5),
            '부정키워드': Counter(neg_found).most_common(5),
        }
    
    @staticmethod
    def analyze_topics(texts):
        """주제별 동향 맥락 분석"""
        topic_results = {}
        for topic, keywords in DataAnalyzer.TOPIC_GROUPS.items():
            count = 0
            examples = []
            for text in texts:
                text_str = str(text)
                if any(kw in text_str for kw in keywords):
                    count += 1
                    if len(examples) < 2:
                        examples.append(text_str[:50])
            if count > 0:
                topic_results[topic] = {'건수': count, '예시': examples}
        return topic_results
    
    @staticmethod
    def generate_context_summary(texts, platform_name):
        """동향 맥락 요약 문장 생성"""
        sentiment = DataAnalyzer.analyze_sentiment(texts)
        topics = DataAnalyzer.analyze_topics(texts)
        keywords = DataAnalyzer.extract_keywords(texts, 5)
        
        summaries = []
        
        # 감성 기반 요약
        pos_ratio = sentiment['긍정비율_num']
        neg_ratio = sentiment['부정비율_num']
        
        if pos_ratio > 50:
            top_pos = sentiment['긍정키워드'][0][0] if sentiment['긍정키워드'] else '긍정적'
            summaries.append(f"전반적으로 '{top_pos}' 등 긍정적 반응이 우세합니다.")
        elif neg_ratio > 40:
            top_neg = sentiment['부정키워드'][0][0] if sentiment['부정키워드'] else '부정적'
            summaries.append(f"'{top_neg}' 등 부정적 의견이 다수 발견됩니다.")
        
        # 주제별 요약
        sorted_topics = sorted(topics.items(), key=lambda x: x[1]['건수'], reverse=True)
        for topic, data in sorted_topics[:2]:
            if data['건수'] >= 3:
                summaries.append(f"'{topic}' 관련 언급이 {data['건수']}건으로 주요 화두입니다.")
        
        # 키워드 기반 요약
        if keywords:
            top_kw = keywords[0][0]
            summaries.append(f"가장 많이 언급된 키워드는 '{top_kw}'입니다.")
        
        return summaries if summaries else ["특이사항 없음"]
    
    @staticmethod
    def summarize_youtube(videos, comments):
        if not videos:
            return {}
        top5 = sorted(videos, key=lambda x: x.get('조회수_raw', 0), reverse=True)[:5]
        for v in top5:
            v_comments = [c['댓글'] for c in comments if c['video_id'] == v['video_id']]
            if v_comments:
                kw = DataAnalyzer.extract_keywords(v_comments, 5)
                v['댓글_키워드'] = ', '.join([f"{k}({c})" for k, c in kw])
        
        all_texts = [c['댓글'] for c in comments]
        dates = [v['업로드일'] for v in videos]
        return {
            'TOP5_영상': top5, '총_영상수': len(videos), '총_댓글수': len(comments),
            '감성분석': DataAnalyzer.analyze_sentiment(all_texts),
            '주요_키워드': DataAnalyzer.extract_keywords(all_texts, 10),
            '주제분석': DataAnalyzer.analyze_topics(all_texts),
            '동향요약': DataAnalyzer.generate_context_summary(all_texts, '유튜브'),
            '분석기간': f"{min(dates)} ~ {max(dates)}" if dates else "N/A"
        }
    
    @staticmethod
    def summarize_dc(posts):
        if not posts:
            return {}
        texts = [p['제목'] + ' ' + p['본문'] for p in posts]
        dates = [p['작성일'][:10] for p in posts if p.get('작성일')]
        return {
            '총_게시글수': len(posts),
            '감성분석': DataAnalyzer.analyze_sentiment(texts),
            '주요_키워드': DataAnalyzer.extract_keywords(texts, 10),
            '주제분석': DataAnalyzer.analyze_topics(texts),
            '동향요약': DataAnalyzer.generate_context_summary(texts, '디시인사이드'),
            '분석기간': f"{min(dates)} ~ {max(dates)}" if dates else "최근 7일"
        }
    
    @staticmethod
    def summarize_reviews(reviews, platform):
        if not reviews:
            return {}
        df = pd.DataFrame(reviews)
        texts = df['내용'].tolist() if '내용' in df else []
        dates = df['작성일'].tolist() if '작성일' in df else []
        rating_dist = df['평점'].astype(float).value_counts().sort_index().to_dict() if '평점' in df else {}
        
        return {
            '총_리뷰수': len(reviews),
            '평균_평점': df['평점'].astype(float).mean() if '평점' in df else 0,
            '평점분포': rating_dist,
            '감성분석': DataAnalyzer.analyze_sentiment(texts),
            '주요_키워드': DataAnalyzer.extract_keywords(texts, 10),
            '주제분석': DataAnalyzer.analyze_topics(texts),
            '동향요약': DataAnalyzer.generate_context_summary(texts, platform),
            '분석기간': f"{min(dates)} ~ {max(dates)}" if dates else "N/A"
        }


# ============================================================
# 텍스트 막대 인포그래픽 생성
# ============================================================
def create_text_bar(pos_ratio, neg_ratio, neu_ratio, width=20):
    """텍스트 기반 막대 그래프 생성"""
    pos_len = int(pos_ratio / 100 * width)
    neg_len = int(neg_ratio / 100 * width)
    neu_len = width - pos_len - neg_len
    
    bar = '█' * pos_len + '▒' * neu_len + '░' * neg_len
    return f"[{bar}] 긍정{pos_ratio:.0f}% 중립{neu_ratio:.0f}% 부정{neg_ratio:.0f}%"


# ============================================================
# 엑셀 대시보드 생성기 (개선)
# ============================================================
class ExcelGenerator:
    def __init__(self, keyword):
        self.keyword = keyword
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def generate(self, yt_videos, yt_comments, dc_posts, appstore, playstore,
                 yt_summary, dc_summary, app_summary, play_summary, ai_report=None):
        filename = f"{self.keyword}_동향분석_{self.timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            self._create_dashboard(writer, yt_summary, dc_summary, app_summary, play_summary, ai_report)
            
            if yt_videos:
                cols = ['순위', '영상 제목', '채널명', '조회수', '좋아요', '댓글수', '업로드일', '영상 설명', '썸네일', '링크']
                df = pd.DataFrame(yt_videos)
                df = df[[c for c in cols if c in df.columns]]
                df.to_excel(writer, sheet_name='유튜브_영상', index=False)
            if yt_comments:
                pd.DataFrame(yt_comments).drop(['video_id'], axis=1, errors='ignore').to_excel(writer, sheet_name='유튜브_댓글', index=False)
            if dc_posts:
                pd.DataFrame(dc_posts).to_excel(writer, sheet_name='디시인사이드', index=False)
            if appstore:
                pd.DataFrame(appstore).drop(['review_id'], axis=1, errors='ignore').to_excel(writer, sheet_name='앱스토어_리뷰', index=False)
            if playstore:
                pd.DataFrame(playstore).to_excel(writer, sheet_name='플레이스토어_리뷰', index=False)
        
        return filename
    
    def _create_dashboard(self, writer, yt, dc, app, play, ai_report=None):
        rows = []
        
        # ═══ 헤더 ═══
        rows.append({'A': f'📊 {self.keyword} 동향 분석 리포트', 'B': '', 'C': ''})
        rows.append({'A': f'생성일시: {datetime.now().strftime("%Y-%m-%d %H:%M")}', 'B': '', 'C': ''})
        rows.append({'A': '', 'B': '', 'C': ''})
        
        # ═══ AI 보고서 (있는 경우) ═══
        if ai_report:
            rows.append({'A': '═' * 70, 'B': '', 'C': ''})
            rows.append({'A': '🤖 AI 동향 분석 보고서', 'B': '', 'C': ''})
            rows.append({'A': '═' * 70, 'B': '', 'C': ''})
            for line in ai_report.split('\n'):
                rows.append({'A': line, 'B': '', 'C': ''})
            rows.append({'A': '', 'B': '', 'C': ''})
        
        # ═══ 1. 전체 오버뷰 ═══
        rows.append({'A': '═' * 70, 'B': '', 'C': ''})
        rows.append({'A': '📈 전체 오버뷰', 'B': '', 'C': ''})
        rows.append({'A': '═' * 70, 'B': '', 'C': ''})
        rows.append({'A': '플랫폼', 'B': '분석 건수', 'C': '분석 기간'})
        rows.append({'A': '─' * 70, 'B': '', 'C': ''})
        
        total_count = 0
        if yt:
            cnt = yt.get('총_영상수', 0) + yt.get('총_댓글수', 0)
            total_count += cnt
            rows.append({'A': '📺 유튜브', 'B': f"영상 {yt.get('총_영상수', 0)}개 + 댓글 {yt.get('총_댓글수', 0)}개", 'C': yt.get('분석기간', 'N/A')})
        if dc:
            total_count += dc.get('총_게시글수', 0)
            rows.append({'A': '💬 디시인사이드', 'B': f"게시글 {dc.get('총_게시글수', 0)}개", 'C': dc.get('분석기간', 'N/A')})
        if app:
            total_count += app.get('총_리뷰수', 0)
            rows.append({'A': '🍎 앱스토어', 'B': f"리뷰 {app.get('총_리뷰수', 0)}개 (평점 {app.get('평균_평점', 0):.1f}점)", 'C': app.get('분석기간', 'N/A')})
        if play:
            total_count += play.get('총_리뷰수', 0)
            rows.append({'A': '🤖 플레이스토어', 'B': f"리뷰 {play.get('총_리뷰수', 0)}개 (평점 {play.get('평균_평점', 0):.1f}점)", 'C': play.get('분석기간', 'N/A')})
        
        rows.append({'A': '─' * 70, 'B': '', 'C': ''})
        rows.append({'A': f'📊 총 분석 데이터: {total_count:,}건', 'B': '', 'C': ''})
        rows.append({'A': '', 'B': '', 'C': ''})
        
        # ═══ 2. 플랫폼별 감성 분석 ═══
        rows.append({'A': '═' * 70, 'B': '', 'C': ''})
        rows.append({'A': '🎯 플랫폼별 감성 분석 (긍정/부정 비율)', 'B': '', 'C': ''})
        rows.append({'A': '═' * 70, 'B': '', 'C': ''})
        
        platforms = [('📺 유튜브', yt), ('💬 디시인사이드', dc), ('🍎 앱스토어', app), ('🤖 플레이스토어', play)]
        
        for pname, summary in platforms:
            if summary and '감성분석' in summary:
                s = summary['감성분석']
                pos = s.get('긍정비율_num', 0)
                neg = s.get('부정비율_num', 0)
                neu = 100 - pos - neg
                
                bar = create_text_bar(pos, neg, neu)
                rows.append({'A': f'{pname}', 'B': '', 'C': ''})
                rows.append({'A': f'  {bar}', 'B': '', 'C': ''})
                
                # 주요 키워드
                if summary.get('주요_키워드'):
                    kw = ', '.join([f"{k}({c})" for k, c in summary['주요_키워드'][:5]])
                    rows.append({'A': f'  주요 키워드: {kw}', 'B': '', 'C': ''})
                
                rows.append({'A': '', 'B': '', 'C': ''})
        
        # 저장
        df = pd.DataFrame(rows)
        df.columns = ['항목', '내용', '비고']
        df.to_excel(writer, sheet_name='📊 대시보드', index=False)


# ============================================================
# Raw 데이터 저장기
# ============================================================
class RawDataSaver:
    def __init__(self, keyword, timestamp):
        self.keyword = keyword
        self.folder = f"{keyword}_raw_{timestamp}"
    
    def save_all(self, yt_videos, yt_comments, dc_posts, appstore, playstore, live_chats=None):
        if not os.path.exists(self.folder):
            os.makedirs(self.folder)
        saved = []
        for name, data in [('유튜브_영상', yt_videos), ('유튜브_댓글', yt_comments),
                           ('디시인사이드', dc_posts), ('앱스토어', appstore),
                           ('플레이스토어', playstore), ('라이브채팅', live_chats)]:
            if data:
                f = f"{self.folder}/{self.keyword}_{name}_raw.csv"
                pd.DataFrame(data).to_csv(f, index=False, encoding='utf-8-sig')
                saved.append(f)
        return self.folder, saved


# ============================================================
# 영상 분석 엑셀 생성기 (세로 레포트 형식)
# ============================================================
class VideoAnalysisExcelGenerator:
    """영상별 세로 레포트 형식 엑셀 생성"""
    
    def generate(self, videos, comments, filename):
        """세로 레포트 형식으로 엑셀 생성"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "📊 영상별 분석 리포트"
            
            # 스타일 정의
            title_font = Font(size=16, bold=True, color="FFFFFF")
            header_font = Font(size=12, bold=True, color="FFFFFF")
            section_font = Font(size=11, bold=True)
            normal_font = Font(size=10)
            
            title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            section_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            row = 1
            
            # 메인 타이틀
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = "📊 유튜브 영상 분석 리포트"
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 30
            row += 1
            
            ws[f'A{row}'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            row += 2
            
            # 각 영상별 레포트
            for idx, video in enumerate(videos, 1):
                # 영상 헤더
                ws.merge_cells(f'A{row}:F{row}')
                cell = ws[f'A{row}']
                title_text = video.get('영상 제목', 'N/A')
                if len(title_text) > 60:
                    title_text = title_text[:60] + "..."
                cell.value = f"═══ 영상 {idx}: {title_text} ═══"
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[row].height = 25
                row += 2
                
                # 📺 기본 정보 섹션
                ws.merge_cells(f'A{row}:F{row}')
                cell = ws[f'A{row}']
                cell.value = "📺 기본 정보"
                cell.font = section_font
                cell.fill = section_fill
                ws.row_dimensions[row].height = 22
                row += 1
                
                info_items = [
                    ('영상 제목', video.get('영상 제목', 'N/A')),
                    ('채널명', video.get('채널명', 'N/A')),
                    ('조회수', video.get('조회수', 'N/A')),
                    ('좋아요', video.get('좋아요', 'N/A')),
                    ('댓글수', video.get('댓글수', 'N/A')),
                    ('업로드일', video.get('업로드일', 'N/A')),
                    ('링크', video.get('링크', 'N/A')),
                ]
                
                for i, (label, value) in enumerate(info_items):
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'A{row}'].border = thin_border
                    ws.merge_cells(f'B{row}:F{row}')
                    ws[f'B{row}'] = str(value)
                    ws[f'B{row}'].border = thin_border
                    if i % 2 == 1:
                        ws[f'A{row}'].fill = alt_fill
                        ws[f'B{row}'].fill = alt_fill
                    row += 1
                row += 1
                
                # 📝 AI 요약 섹션 (있는 경우)
                if video.get('AI 요약'):
                    ws.merge_cells(f'A{row}:F{row}')
                    cell = ws[f'A{row}']
                    cell.value = "📝 AI 요약"
                    cell.font = section_font
                    cell.fill = section_fill
                    ws.row_dimensions[row].height = 22
                    row += 1
                    
                    summary_lines = video['AI 요약'].split('\n')
                    for line in summary_lines:
                        if line.strip():
                            ws.merge_cells(f'A{row}:F{row}')
                            ws[f'A{row}'] = line
                            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                            row += 1
                    row += 1
                
                # 📊 댓글 감성 분석 섹션 (있는 경우)
                if video.get('댓글 감성 분석'):
                    # 정량적 분석
                    ws.merge_cells(f'A{row}:F{row}')
                    cell = ws[f'A{row}']
                    cell.value = "📊 댓글 감성 분석"
                    cell.font = section_font
                    cell.fill = section_fill
                    ws.row_dimensions[row].height = 22
                    row += 1
                    
                    sentiment_text = video['댓글 감성 분석']
                    sentiment_lines = sentiment_text.split('\n')
                    for line in sentiment_lines:
                        if line.strip():
                            ws.merge_cells(f'A{row}:F{row}')
                            ws[f'A{row}'] = line
                            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                            row += 1
                    row += 1
                
                # 구분선
                row += 2
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            
            # 전체 댓글 시트 추가
            if comments:
                ws_comments = wb.create_sheet("전체 댓글")
                df_comments = pd.DataFrame(comments)
                
                # 헤더
                for col_idx, col_name in enumerate(df_comments.columns, 1):
                    cell = ws_comments.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # 데이터
                for row_idx, row_data in enumerate(df_comments.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws_comments.cell(row=row_idx, column=col_idx, value=value)
                
                # 열 너비 자동 조정
                for col_idx, col_name in enumerate(df_comments.columns, 1):
                    ws_comments.column_dimensions[get_column_letter(col_idx)].width = 20
            
            wb.save(filename)
            return filename
            
        except Exception as e:
            # 스타일링 없이 기본 방식으로 저장
            return self._generate_basic(videos, comments, filename)
    
    def _generate_basic(self, videos, comments, filename):
        """스타일링 없이 기본 엑셀 생성 (fallback)"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            rows = []
            for idx, video in enumerate(videos, 1):
                rows.append({'항목': f'═══ 영상 {idx} ═══', '내용': video.get('영상 제목', '')})
                rows.append({'항목': '채널명', '내용': video.get('채널명', '')})
                rows.append({'항목': '조회수', '내용': video.get('조회수', '')})
                rows.append({'항목': '좋아요', '내용': video.get('좋아요', '')})
                rows.append({'항목': '업로드일', '내용': video.get('업로드일', '')})
                rows.append({'항목': '링크', '내용': video.get('링크', '')})
                rows.append({'항목': '', '내용': ''})
                
                if video.get('AI 요약'):
                    rows.append({'항목': '📝 AI 요약', '내용': ''})
                    for line in video['AI 요약'].split('\n'):
                        if line.strip():
                            rows.append({'항목': '', '내용': line})
                    rows.append({'항목': '', '내용': ''})
                
                if video.get('댓글 감성 분석'):
                    rows.append({'항목': '📊 댓글 감성 분석', '내용': ''})
                    for line in video['댓글 감성 분석'].split('\n'):
                        if line.strip():
                            rows.append({'항목': '', '내용': line})
                    rows.append({'항목': '', '내용': ''})
                
                rows.append({'항목': '', '내용': ''})
            
            pd.DataFrame(rows).to_excel(writer, sheet_name='영상별 분석', index=False)
            
            if comments:
                pd.DataFrame(comments).to_excel(writer, sheet_name='전체 댓글', index=False)
        
        return filename


# ============================================================
# GUI 애플리케이션
# ============================================================
class AnalyzerApp:
    # 호랑이 이미지 (Base64 인코딩)
    TIGER_IMAGE = """
iVBORw0KGgoAAAANSUhEUgAAAGQAAABBCAIAAACo4ZaGAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAA3W0lEQVR4nDW7Wa+v2XHeV7Wud/qPezz7zD2f02Q32WxxaM6kKMuybHmSYiNAYtgXATI4QHyR3ATxXb5AEicxnAQwhAwwYsmGrNgxLckmRVKcusmeTvfpPtPe++zxP7/TGqoqF418g4XCWk89T61f4d/9z/+jalBa7XanU+3c+cXl6cnT9z/4aLVs+r61xlzbnuxtVdtbo+H2ddEuUmLitqnXq03o29EwX22aJ0fngPHq1uDTz16/+9qz05v7k/FeleWxbpeXc63saDQGpxOrEH3mjGKBwOPp2OaOY/IpCrMxtldaG6RAoFMK3q+70A+qCrRiIKMQtGKtBIlFhAW1BhHpe0qJNYIIBCJBEZDERMQ+JWYiQJsJEkVOAVFrBhIRSiH5FJIiVE7TJxGZJcSUOCIzIyolzMKSCDlFEgKlNAoTICALkLCINECgYwAhpZICAkZOPRNRbBMwMxAYpS0asgZJQGl0ygDoEHzXxxC9j6JYjLNaWWWMyhwJozJd33uOgtT1IQoYpXcHBYKywH3fkwigMggaJCRCYoGUhFiZLEuJkDj6yCKojFGKiQFRgGJIRCTMRITCBIJKsYhGDSwsAJpJSAAkCYMARwISRkRhZYgREROzsCAgsyABABBLYhYQECBgJmLBRMlHjiQJWJABWIBYBBCYRBQQgAJKACCIqAAEEJhIKaUAxCcWTkmSJEpEIikJC4lEFkCjlUYWQEYlKJhICAVCCj4liokEOQkjBCIUSSmFFJmTNkpAQAAJUSxqAYQIIAAEiCgQ4X8SIU6MSEo0CQqIYhSlGFEk+BQ5MlJiIIrMCsAzKaYcSYMBksjEwhKJQvKJSRMpRYiYmIQVsgAyozCLAgJQEYREkAQYOSUm4L7vgUWYmUUAWFB8IkqRU0hMIjFGgaRBeUoJgJOgAKIwg7BIAmIiJpKUCFiiMBOrT5R8TKJJGIgBBAmJATgJMjALImgGJYkJUCekJJI4MQiwICASSBQRZJQkqBWQCIhAYhAETgKowCdKSQhYCQigSBSNLARAiCwCFEQYQZgYAREgEaCiiCjAwgJMEpkAJUoSxsiskZmFKAmLBtAigIKCEsUxBRZUIChKkChIRiOFkVEUBCaKwhIJEiWBJAHJkCASJBFKhCmyMAIJAEAmCSSRAKKiBNYCoFIKFVliSQJJEqIGYS0EwsDAQKlLjIJMAIzIgIKIoJFJhASJQVgIBBQwECIyIJJmcQSsFAMyBxKgKJCYAJF1AZJEECETICAJJbZIwIigJIpBCgwEhBRAKAArBkrIgIggCBoUJAGIBKQAAJGAELVCRkJBRBISdgiMQoLakLAgCAMgK0RAYkANAIAIjIgEJKQENIBiQRBUAiQkKKARCJEJIYqQsKKEHIkJlFYiyBKJUIkA0icoVAAFmAERkJARAJAIEBGIEZEBFQCDAAKoTwJD/ERfUjKJNMAnzhCZmBkADAAzQEIkgAQgqBQSUkIJIRijWbNEQEKFSMhAooRFI4MCZiJQBMAgJAkBFSASCwAzkQAgIyuhJBw4aRIFgAqBFDICKwYCEmIEJBRRAsCI/D8JJkBhAALgJIgoCQiAGREIUEAApCdhDYgoCsCQBEqMIBqJGIGQACUlIQGlFAMCoSQkpURQAhIxMgAlZgCFSjMTCCAggQJETqSQlCAJCqByiAJKBIUogqAQCqYkCJCQQSEDQCKlBNAnIBJFiABBRCEQKgUICIAEgpCEAKwAMMQAlAgJQJBBABRAEomOiBE1o0IAIAWYEIEBOAEJmAQGFQghIwEoYQABQkQGoAgRQIFoAoiIgMIiLMAQkYgZJAkDkCADKhEE0AHQJkTUQggCIAAIQIIALKBBkMAQaEQBRAUMKAwJKbIRjIDABABICATIMEIWhYICIgIAxEAIwowkwoIgSgFACIgiACjAoAARtQArAgCFiFoIGQUxEQkIIwgqBhBACQJJQEQRaAIhEiEGAUFIAiygjAQhUkgRUWnKiCR6IkbhRBwBEiIpDZ+QS0qglCZAJBFBICIAxgSKREACRpQkkkSAQRJoEkJhDQJAIEFSkkRjYiYUAJYoKIBJhJAiYCIQQQFIEkUACREToUYACQQAICQIMTIKKCWIJJASaQESFJCkAKRAEAKkJAiCwgIEhAwsiEgICIAaEVCARYQIQAGigAggkgKgFJKIMBIJZogICZQiIkAUBCZkEEJBSEQqiqAAAokoBibQiIwYiRMBAmECUEiCBABCKIE4kWJgBImklAAwIQmAgJAiMCJAAgiAqAAIEBISECBACZETChMIigJQAEJKgQiDCIqAKIUECCJJQAMoBCBKyIiCKIQghEAQSCWKSAIkIECCCAIMwCCAiCRIRBCJNICyGkAjRk7akBATEQiIBhFhUkgJUREQMRKCUpoAFIhOkjgJJSKFlCQJEhKgIiFAAIBEjEggRIgIAAoRJBEoVMIJmZiVBgZhREDNwEgISIIMJEyCQoKIwEmYJEICBIigCAQ5JbWJKRESIkUiBgACJSCAyAqFAFgEBJKQJGYBUAAaBAgQEqKARIpECZAiiREYhD6hgUmBQCQJIgIACCEAQxJOQMSIpEESKWREFiEkFKUEkJNoBkIJLAKE+ImKBEARiYBoAQ2oQRKgsBBHUKISRwAJiJLAJxWLEiWhCESEhCwiiQWBhRWgiEQgIUYCEUYCIWAUQEQkINaIimIkJCAhYAEiFgIhgEgQQQIiCYsmJIqKSAgVJxQBEiJERAJkBYwaCJBJBBAwIoIEECJIQIkQhQCQEqIACCAjSiQQDAAAqBQkAU4iIIkVE4A2KnIyGgEgMRMJgwCiYkBCFA0IiRkTAKJCApBERAKEhIgEIAzEjABCwMJMgkhakFl8YAZEIWERhYQERIgkhIgKCJFYAAURADnxJz3s/3dKAAJIggkIBZmIARQgqAQSgQBQQIgQBYkkMaICYQBCBOZEgAQEIqgAUJIgEhAgkSAgIIFIEkYARGUUB0kJICFqRCWJAFFRIgABQBAAEGVANJBCIBQBSYSEEIgBBJhQkEEASYhAEEVYISghRAIUECJBRmIkRmFByJgAIQpqQgEQBEASTkgMQqiYCAAZBRBBCKNCoMjIoISBOCIRAGoGjEIKQAAUggCyMBBrAgImIoOUBAhBARALIxAQsoAQATBBQpTEoJglChIHYA0AQqQQEohCiQRAgKBBqJGAiAhERCELJCYUgEiaQIQBhAACCYowJEiCCYSQCSASJCGVQJSIMAmJMGgARAREJRoUACIhooAgMAsKJACNDEAgiIKMn0R8ACJQQCCJEYk4EJAAAxCQQCASJCJFpFgRI6M2KRBKEkEUYkQCFhaIhEmEkZE1cFIkBAAC+pMcTyCIAgxCggREQooQgVAhKEaBFAARmUg+6cVCYFiYlDISJ1YRAREIBAAAICQICP8/ik4gAYUoKCgiJEIIyEhKAECYAACECSARCDOBNgjCjARCQkgJhEEhCCIJJAAGECUSCRQyICABCYMGYVBCAswIDAqRWQgJBYEUJFAkSMIgAghICJoVAAAqQAJBSCTACIQoLAQAGgAASIQggQgwIqIAAgoIMAogaQJAAUJARC0QhQQSCQGxAAInICRWwkyILACCJETACECAAizCwowMCoiEQRCYABiYgBEAkISEUQMxAQAkZGIBIWJBJJSQkAQRgFCYgZkTCwoJAgiJEIAGoigSEjhBFqAoAETICpQAChKQCIMAKgIBYU6oSBAgAhJAAowJBAAEBBQwACACAQCQYmACTaAIkJREBkJA0CgCSgQQCVAIABgB0BARIAMgggAJCCIiAwAAsAAgECESsKAQkAByCgBAoBAwEAAhAAIigAAKICoEYQBBUgAIKAJERAqEQCMqEGAGJGSJABAZARgICQWFAAi0AGpWCCJAKIpIiAAEAAYZhZVBZEAAEAUCjAAIAEIECIKKSRAIlCBDFCKdCEgJCUggChFCIiEBRRCBEFCYEBEBBYDpE2ISJCaKhJJAAUVBBRIJCRUqBYwkLAqAE1FCJCZUJEgJFIgAoARGICFmRYAsBBoIQACANSYMhAwALJRAIRICRQQABAwIqD1xZEJiEI1aJREFQIgIoIkIECAxsCAqAU4IoIyJABpAIyQEyowIQEKsEiRCAhECpABa8SfSJJIEAQxAFKAoiIiAYFgpkkiE/D8pJtEkIsDAGhKiIBCJIhIEJCRAJYgiCCzCQoBKAyISIgKAgASEyMzASGwYEJWAICGAoCIFgESoUABRhBFBIQggCitIwlqYGREpIasIjBSAgRNHQGIkRmEkISBCAEQSZiEAQkJCRQgMhIggFIUjCCSmJJEIABQKCqEoREEBBYAJIKKApAQkIMCKIaFGlEhAggxAIqCJkogAAoBCBkACApTIwoKgBIQFkggiaoYkoFCAIIkAIgqQIAiIIiAQAIVKgCMQshKBRCSIKISECEJMAIggSCDADEoDAqFEIBEUgIhagBMggpLEIAAgQAAKIDJIChQJNBIKkyKITJEJQYEIKgAiBkBSwiAkFBNDjEopBQQAhECKQAgBBFCAGIEFRERERJgSJYCEqJMkAGBgYGJgEkggBIxCBIwsIokQNRAICAOiAKICoJSIWYAESIQBGJCBEgAlAgZCYiQhFBBJCBIEhQAoohCVMBEIICqhkFAAAFgJowIEUsTICCiCyIhCQMKJkIkAxSgyIACREBIqiYJCiEoxABMwMoAkQaAEygKCEOmEogEhEQsII0lKggQIoBAAgVAJgCKBJEwqJoGEBgARlREEBgUiCpAIABQAERCBAhVARIEIxEggghKQCIEBBAQEBZAIkDUCgoggICBFJEiEACiJAJIiAiBEECABIMGIJKgAkT5h54ggiOoTLMAgKBqQgAQIiRVKQhEBpESkERIwKEJBZCIQFpKIIAiggBmAADQC0Se6VhJREowMgPxJMCcECQgRgBAyaEGBCAqImBARCZBJkIE1ISsRJAAQIgYGZOFEHBgYCEAiIBIwaCAAQEQFCIQMgJSAEFCyiGgJApIIAIBJBAhREhAwISIRkgACEjKJEIj6BEsAJU4RSEAhACQiYSQAIkBhImAAYRQhACJCYWFGQQQSIQCMqBRSEEIAJJCELIzITIqQwIAmQGLFgJqJCRlICCIKMyoC0iAJkBiREEUYgJIksETEyIkEBYEQCIgIgQQBMH0CMZTICRiJkJBAKSRFISGjMCAQMBGDcBIhAiAQ0ASAHOkTrgiJRABIKREhIAgLAwMoCCCAzEQoIoiIiCAggggEBMiaASkJEyKDYiEAAgYUQBFgQlRAkAgEGBGQEQAQERH/fxQdGQWIQGlBTACsSCkkAEIOgAACoIAFCBARFDCyJIQEIECEqJAIIbFCEQAREQIGJmIUYAUogAKAQESREgKwEogCCABIKIKsQAgABEABIhIgECQhRCYERhQkUkAIQKSUROIkAsmApEQixERCREiCQKgYWBQJEZAASGIhFhYQIRIEJCIhYERSCCCqCAiJRAAxkSACIgsICwugsCYRSQKCSIBAIIKAIEKEQMikQASZhQAIhIgkJkEhISEESEmAEYQEUBKIEghKokgiwoJITICAIpBQNAAhAICARBJmZgIRFIUkIhI+AamAEBRCAoWgBAFQIiILIWpBUpoliSiEjxWCAiAkRAUgrCAJkiAAIioGJKBEDEhCyAgAEhQKyP8PJkoiAErTJwImiRIBJEQAEE1IKJhIAahECAhCkCgBgwgCASCQJkAmFCAEAgCJDCIMAMKCCIIIIIyCnEgZACAERgAiVBKBkAAJBBAAIFIUQEQAIBQCAAJCIEJEIEACABAGAsREiAoAAQSJ/j8U/S+IRQW3ycFvjgAAAABJRU5ErkJggg==
"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🐯 동향 수집하는 호랑이 v2.3")
        self.root.geometry("750x950")
        self.root.resizable(False, False)
        self.result_file = None
        self.live_stop_event = None
        self.live_chats = []
        self.video_link_entries = []  # 유튜브 영상 분석용 링크 입력 칸들
        self.video_analysis_result = None
        self.create_widgets()
    
    def create_widgets(self):
        # 타이틀 프레임 (이미지 + 제목)
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # 호랑이 이미지 로드
        try:
            import base64
            from io import BytesIO
            from PIL import Image, ImageTk
            
            img_data = base64.b64decode(self.TIGER_IMAGE.strip())
            img = Image.open(BytesIO(img_data))
            self.tiger_photo = ImageTk.PhotoImage(img)
            
            # 이미지 라벨
            img_label = ttk.Label(title_frame, image=self.tiger_photo)
            img_label.pack(side=tk.LEFT, padx=(0, 10))
        except:
            pass  # 이미지 로드 실패 시 무시
        
        # 타이틀 텍스트
        title_text = ttk.Label(title_frame, text="🐯 동향 수집하는 호랑이 v2.3", font=('맑은 고딕', 18, 'bold'))
        title_text.pack(side=tk.LEFT)
        
        # 부제목
        subtitle = ttk.Label(title_frame, text="YouTube · 앱스토어 · 플레이스토어 · 디시인사이드", 
                            font=('맑은 고딕', 9), foreground='gray')
        subtitle.pack(side=tk.LEFT, padx=(10, 0))
        
        # 노트북 (탭)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))
        
        # 탭 1: 동향수집기
        self.analysis_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_tab, text="📊 키워드 동향")
        self.create_analysis_tab()
        
        # 탭 2: 유튜브 영상 분석
        self.video_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.video_tab, text="📺 유튜브 영상 분석")
        self.create_video_tab()
        
        # 탭 3: 라이브 채팅
        self.live_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.live_tab, text="🔴 라이브 채팅")
        self.create_live_tab()
    
    def create_analysis_tab(self):
        main = ttk.Frame(self.analysis_tab, padding="10")
        main.pack(fill=tk.BOTH, expand=True)
        
        # 기본 설정
        basic = ttk.LabelFrame(main, text="📌 기본 설정", padding="10")
        basic.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(basic, text="분석 키워드:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.keyword_var = tk.StringVar(value="메이플키우기")
        ttk.Entry(basic, textvariable=self.keyword_var, width=40).grid(row=0, column=1, columnspan=2, sticky=tk.W)
        ttk.Label(basic, text="(유튜브·디시 키워드 검색 사용 시에만 필수)", font=('맑은 고딕', 8), foreground='gray').grid(row=1, column=1, sticky=tk.W)
        
        ttk.Label(basic, text="YouTube API 키:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.api_key_var = tk.StringVar(value="AIzaSyDC2YFzCPEuSHeaX3d2ZNc-z6GmzFO-ROg")
        ttk.Entry(basic, textvariable=self.api_key_var, width=40).grid(row=2, column=1, columnspan=2, sticky=tk.W)
        ttk.Label(basic, text="※ 기본 키 제공됨", font=('맑은 고딕', 8), foreground='gray').grid(row=3, column=1, sticky=tk.W)
        
        # 앱 설정
        app_frame = ttk.LabelFrame(main, text="📱 스토어 설정", padding="10")
        app_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(app_frame, text="앱스토어 ID:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.appstore_var = tk.StringVar(value="6739616715")
        ttk.Entry(app_frame, textvariable=self.appstore_var, width=25).grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(app_frame, text="플레이스토어 ID:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.playstore_var = tk.StringVar(value="com.nexon.ma")
        ttk.Entry(app_frame, textvariable=self.playstore_var, width=25).grid(row=1, column=1, sticky=tk.W)
        
        # 디시 설정 (키워드 검색 / 갤러리 수집 분리)
        dc_frame = ttk.LabelFrame(main, text="💬 디시인사이드 설정", padding="10")
        dc_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.dc_keyword_check = tk.BooleanVar(value=True)
        ttk.Checkbutton(dc_frame, text="키워드 검색 수집 (분석 키워드로 전체 검색)", variable=self.dc_keyword_check).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=2)
        
        self.dc_gallery_check = tk.BooleanVar(value=True)
        ttk.Checkbutton(dc_frame, text="갤러리 글 수집 (아래 갤러리 ID 기준)", variable=self.dc_gallery_check).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=2)
        
        ttk.Label(dc_frame, text="갤러리 ID:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.gallery_var = tk.StringVar(value="maplerpg")
        ttk.Entry(dc_frame, textvariable=self.gallery_var, width=20).grid(row=2, column=1, sticky=tk.W)
        
        self.is_minor_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(dc_frame, text="마이너 갤러리", variable=self.is_minor_var).grid(row=2, column=2, padx=10)
        
        # 수집 옵션
        opt_frame = ttk.LabelFrame(main, text="⚙️ 수집 옵션", padding="10")
        opt_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 체크박스 + 드롭다운
        self.yt_check = tk.BooleanVar(value=True)
        self.dc_check = tk.BooleanVar(value=True)
        self.app_check = tk.BooleanVar(value=True)
        self.play_check = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(opt_frame, text="📺 유튜브", variable=self.yt_check).grid(row=0, column=0, sticky=tk.W)
        ttk.Label(opt_frame, text="영상 수:").grid(row=0, column=1, padx=(10, 5))
        self.yt_count_var = tk.StringVar(value="20")
        yt_combo = ttk.Combobox(opt_frame, textvariable=self.yt_count_var, values=["20", "50", "100"], width=5, state="readonly")
        yt_combo.grid(row=0, column=2)
        
        ttk.Checkbutton(opt_frame, text="💬 디시", variable=self.dc_check).grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Label(opt_frame, text="페이지:").grid(row=1, column=1, padx=(10, 5))
        self.dc_page_var = tk.StringVar(value="10")
        dc_combo = ttk.Combobox(opt_frame, textvariable=self.dc_page_var, values=["3", "5", "10", "50"], width=5, state="readonly")
        dc_combo.grid(row=1, column=2)
        ttk.Label(opt_frame, text="(3p~1분 / 5p~2분 / 10p~4분 / 50p~20분)", font=('맑은 고딕', 8), foreground='gray').grid(row=1, column=3, padx=5)
        
        ttk.Checkbutton(opt_frame, text="🍎 앱스토어", variable=self.app_check).grid(row=2, column=0, sticky=tk.W)
        ttk.Checkbutton(opt_frame, text="🤖 플레이스토어", variable=self.play_check).grid(row=2, column=1, columnspan=2, sticky=tk.W)
        
        ttk.Separator(opt_frame, orient='horizontal').grid(row=3, column=0, columnspan=4, sticky='ew', pady=10)
        self.raw_check = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_frame, text="📁 Raw 데이터 별도 저장 (CSV)", variable=self.raw_check).grid(row=4, column=0, columnspan=2, sticky='w')
        
        # AI 보고서 옵션
        self.ai_report_check = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="🤖 AI 동향 분석 보고서 생성 (Gemini)", variable=self.ai_report_check).grid(row=5, column=0, columnspan=3, sticky='w', pady=(5,0))
        ttk.Label(opt_frame, text="※ Gemini API 키 필요 (유튜브 영상 분석 탭에서 설정)", font=('맑은 고딕', 8), foreground='gray').grid(row=6, column=0, columnspan=4, sticky='w')
        
        # 버튼
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=10)
        
        self.start_btn = ttk.Button(btn_frame, text="🚀 분석 시작", command=self.start_analysis)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.open_btn = ttk.Button(btn_frame, text="📊 대시보드 열기", command=self.open_result, state=tk.DISABLED)
        self.open_btn.pack(side=tk.LEFT, padx=5)
        self.folder_btn = ttk.Button(btn_frame, text="📂 폴더 열기", command=self.open_folder, state=tk.DISABLED)
        self.folder_btn.pack(side=tk.LEFT, padx=5)
        
        # 진행 상황
        prog_frame = ttk.LabelFrame(main, text="📊 진행 상황", padding="10")
        prog_frame.pack(fill=tk.BOTH, expand=True)
        
        self.progress_var = tk.DoubleVar()
        ttk.Progressbar(prog_frame, variable=self.progress_var, maximum=100).pack(fill=tk.X, pady=(0, 10))
        self.log_text = scrolledtext.ScrolledText(prog_frame, height=10, font=('Consolas', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def create_live_tab(self):
        main = ttk.Frame(self.live_tab, padding="10")
        main.pack(fill=tk.BOTH, expand=True)
        
        # 안내 문구
        warn_frame = ttk.Frame(main)
        warn_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(warn_frame, text="⚠️ 라이브 채팅은 '수집 시작' 버튼을 누른 시점부터의 채팅만 수집됩니다.", 
                  foreground='red', font=('맑은 고딕', 9, 'bold')).pack()
        
        # 플랫폼 선택
        platform_frame = ttk.LabelFrame(main, text="📺 플랫폼 선택", padding="10")
        platform_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.live_platform_var = tk.StringVar(value="youtube")
        ttk.Radiobutton(platform_frame, text="유튜브 라이브", variable=self.live_platform_var, value="youtube").pack(side=tk.LEFT, padx=10)
        
        # URL 입력
        url_frame = ttk.LabelFrame(main, text="🔗 라이브 URL", padding="10")
        url_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(url_frame, text="URL:").pack(side=tk.LEFT)
        self.live_url_var = tk.StringVar()
        ttk.Entry(url_frame, textvariable=self.live_url_var, width=60).pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        ttk.Label(url_frame, text="유튜브: youtube.com/watch?v=xxx", 
                  font=('맑은 고딕', 8), foreground='gray').pack(side=tk.BOTTOM, anchor=tk.W, pady=(5, 0))
        
        # 버튼
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=10)
        
        self.live_start_btn = ttk.Button(btn_frame, text="▶ 수집 시작", command=self.start_live_collection)
        self.live_start_btn.pack(side=tk.LEFT, padx=5)
        self.live_stop_btn = ttk.Button(btn_frame, text="⏹ 수집 중지", command=self.stop_live_collection, state=tk.DISABLED)
        self.live_stop_btn.pack(side=tk.LEFT, padx=5)
        self.live_save_btn = ttk.Button(btn_frame, text="💾 CSV 저장", command=self.save_live_chats, state=tk.DISABLED)
        self.live_save_btn.pack(side=tk.LEFT, padx=5)
        
        # 상태
        self.live_status_var = tk.StringVar(value="대기 중...")
        ttk.Label(btn_frame, textvariable=self.live_status_var, foreground='blue').pack(side=tk.LEFT, padx=20)
        
        # 채팅 로그
        log_frame = ttk.LabelFrame(main, text="💬 채팅 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.live_log = scrolledtext.ScrolledText(log_frame, height=15, font=('Consolas', 9))
        self.live_log.pack(fill=tk.BOTH, expand=True)
    
    def create_video_tab(self):
        """유튜브 영상 분석 탭 생성"""
        main = ttk.Frame(self.video_tab, padding="10")
        main.pack(fill=tk.BOTH, expand=True)
        
        # 안내 문구
        info_frame = ttk.Frame(main)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(info_frame, text="📺 유튜브 영상 링크를 입력하면 영상 정보, AI 요약, 전체 댓글을 수집합니다.", 
                  font=('맑은 고딕', 9)).pack(anchor=tk.W)
        ttk.Label(info_frame, text="※ Gemini AI가 영상을 분석하여 요약합니다. (google-generativeai 필요)", 
                  font=('맑은 고딕', 8), foreground='gray').pack(anchor=tk.W)
        
        # API 키 설정
        api_frame = ttk.LabelFrame(main, text="🔑 API 키 설정", padding="10")
        api_frame.pack(fill=tk.X, pady=(0, 10))
        
        # YouTube API 키 안내
        ttk.Label(api_frame, text="※ YouTube API 키는 '키워드 동향' 탭의 키를 사용합니다.", 
                  font=('맑은 고딕', 8), foreground='gray').grid(row=0, column=0, columnspan=3, sticky=tk.W)
        
        # Gemini API 키 입력
        ttk.Label(api_frame, text="Gemini API 키:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.gemini_api_key_var = tk.StringVar(value="AIzaSyCh2MvocCCySAyAUhBGPNj3zl-lT1jKyNE")
        ttk.Entry(api_frame, textvariable=self.gemini_api_key_var, width=45).grid(row=1, column=1, sticky=tk.W, padx=5)
        ttk.Label(api_frame, text="※ 기본 키 제공됨", 
                  font=('맑은 고딕', 8), foreground='gray').grid(row=1, column=2, sticky=tk.W)
        
        # 경고 문구
        ttk.Label(api_frame, text="⚠️ 여러 사용자가 동시에 사용 시 오류가 발생할 수 있습니다. 개인 API 키 사용 권장 (aistudio.google.com)", 
                  font=('맑은 고딕', 8), foreground='red').grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(2,0))
        
        # AI 요약 사용 여부 체크박스
        self.use_gemini_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(api_frame, text="🤖 Gemini AI 영상 요약 사용", 
                       variable=self.use_gemini_var).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(5,0))
        
        # 대댓글 수집 체크박스
        self.include_replies_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(api_frame, text="💬 대댓글도 수집 (시간 증가)", 
                       variable=self.include_replies_var).grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=(2,0))
        
        # 링크 입력 영역 (스크롤 가능)
        links_outer_frame = ttk.LabelFrame(main, text="🔗 유튜브 링크 입력 (최대 20개)", padding="10")
        links_outer_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Canvas + Scrollbar로 스크롤 가능한 영역 생성
        canvas = tk.Canvas(links_outer_frame, height=200)
        scrollbar = ttk.Scrollbar(links_outer_frame, orient="vertical", command=canvas.yview)
        self.links_frame = ttk.Frame(canvas)
        
        self.links_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.links_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 마우스 휠 스크롤 바인딩
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # 링크 추가 버튼 (먼저 생성)
        add_btn_frame = ttk.Frame(main)
        add_btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.add_link_btn = ttk.Button(add_btn_frame, text="+ 링크 추가", command=self.add_link_entry)
        self.add_link_btn.pack(side=tk.LEFT)
        
        self.link_count_label = tk.StringVar(value="(0/20)")
        ttk.Label(add_btn_frame, textvariable=self.link_count_label, foreground='gray').pack(side=tk.LEFT, padx=10)
        
        ttk.Button(add_btn_frame, text="🗑️ 모두 지우기", command=self.clear_all_links).pack(side=tk.RIGHT)
        
        # 기본 링크 입력 칸 3개 생성
        for i in range(3):
            self.add_link_entry()
        
        # 버튼 영역
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=10)
        
        self.video_start_btn = ttk.Button(btn_frame, text="🚀 분석 시작", command=self.start_video_analysis)
        self.video_start_btn.pack(side=tk.LEFT, padx=5)
        
        self.video_open_btn = ttk.Button(btn_frame, text="📊 결과 열기", command=self.open_video_result, state=tk.DISABLED)
        self.video_open_btn.pack(side=tk.LEFT, padx=5)
        
        # 진행 상황
        prog_frame = ttk.LabelFrame(main, text="📊 진행 상황", padding="10")
        prog_frame.pack(fill=tk.BOTH, expand=True)
        
        self.video_progress_var = tk.DoubleVar()
        ttk.Progressbar(prog_frame, variable=self.video_progress_var, maximum=100).pack(fill=tk.X, pady=(0, 10))
        
        self.video_log = scrolledtext.ScrolledText(prog_frame, height=8, font=('Consolas', 9))
        self.video_log.pack(fill=tk.BOTH, expand=True)
    
    def add_link_entry(self):
        """링크 입력 칸 추가"""
        if len(self.video_link_entries) >= 20:
            messagebox.showwarning("알림", "최대 20개까지만 추가할 수 있습니다.")
            return
        
        idx = len(self.video_link_entries) + 1
        
        frame = ttk.Frame(self.links_frame)
        frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(frame, text=f"[{idx}]", width=4).pack(side=tk.LEFT)
        
        entry_var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=entry_var, width=60)
        entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        del_btn = ttk.Button(frame, text="✕", width=3, 
                            command=lambda f=frame, e=(entry_var, frame): self.remove_link_entry(e))
        del_btn.pack(side=tk.RIGHT)
        
        self.video_link_entries.append((entry_var, frame))
        self.update_link_count()
    
    def remove_link_entry(self, entry_tuple):
        """링크 입력 칸 삭제"""
        if len(self.video_link_entries) <= 1:
            messagebox.showwarning("알림", "최소 1개의 입력 칸이 필요합니다.")
            return
        
        entry_var, frame = entry_tuple
        frame.destroy()
        self.video_link_entries.remove(entry_tuple)
        self.renumber_entries()
        self.update_link_count()
    
    def renumber_entries(self):
        """입력 칸 번호 재정렬"""
        for idx, (entry_var, frame) in enumerate(self.video_link_entries, 1):
            for child in frame.winfo_children():
                if isinstance(child, ttk.Label):
                    child.config(text=f"[{idx}]")
                    break
    
    def clear_all_links(self):
        """모든 링크 지우기"""
        for entry_var, frame in self.video_link_entries:
            entry_var.set("")
    
    def update_link_count(self):
        """링크 개수 업데이트"""
        count = len(self.video_link_entries)
        self.link_count_label.set(f"({count}/20)")
        
        if count >= 20:
            self.add_link_btn.config(state=tk.DISABLED)
        else:
            self.add_link_btn.config(state=tk.NORMAL)
    
    def video_log_msg(self, msg):
        """영상 분석 로그 출력"""
        self.video_log.insert(tk.END, f"{msg}\n")
        self.video_log.see(tk.END)
        self.root.update_idletasks()
    
    def extract_video_id(self, url):
        """URL에서 video_id 추출"""
        patterns = [
            r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([a-zA-Z0-9_-]{11})',
            r'([a-zA-Z0-9_-]{11})'
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None
    
    def start_video_analysis(self):
        """영상 분석 시작"""
        # 입력된 링크 수집
        links = []
        for entry_var, frame in self.video_link_entries:
            url = entry_var.get().strip()
            if url:
                video_id = self.extract_video_id(url)
                if video_id:
                    links.append((url, video_id))
                else:
                    messagebox.showerror("오류", f"유효하지 않은 링크입니다:\n{url}")
                    return
        
        if not links:
            messagebox.showerror("오류", "최소 1개의 유튜브 링크를 입력해주세요!")
            return
        
        # YouTube API 키 확인 (동향수집기 탭의 키 사용)
        api_key = self.api_key_var.get().strip()
        if not api_key:
            messagebox.showerror("오류", "키워드 동향 탭에서 YouTube API 키를 입력해주세요!")
            return
        
        # Gemini API 키 확인 (AI 요약 사용 시)
        gemini_key = self.gemini_api_key_var.get().strip()
        use_gemini = self.use_gemini_var.get()
        
        if use_gemini and not gemini_key:
            messagebox.showerror("오류", "Gemini API 키를 입력해주세요!\n(AI 요약을 사용하지 않으려면 체크 해제)")
            return
        
        if use_gemini and not GEMINI_AVAILABLE:
            messagebox.showerror("오류", "google-generativeai 라이브러리가 설치되지 않았습니다.\npip install google-generativeai")
            return
        
        self.video_start_btn.config(state=tk.DISABLED)
        self.video_open_btn.config(state=tk.DISABLED)
        self.video_log.delete(1.0, tk.END)
        self.video_progress_var.set(0)
        
        include_replies = self.include_replies_var.get()
        threading.Thread(target=self.run_video_analysis, args=(links, api_key, gemini_key, use_gemini, include_replies), daemon=True).start()
    
    def run_video_analysis(self, links, api_key, gemini_key, use_gemini, include_replies):
        """영상 분석 실행"""
        try:
            self.video_log_msg(f"🎬 {len(links)}개 영상 분석 시작!")
            if use_gemini:
                self.video_log_msg("🤖 Gemini AI 영상 요약 활성화")
            if include_replies:
                self.video_log_msg("💬 대댓글 수집 활성화 (시간이 더 걸릴 수 있습니다)")
                self.video_log_msg("🤖 Gemini AI 영상 요약 활성화")
            self.video_log_msg("=" * 50)
            
            yt = YouTubeCrawler(api_key)
            
            # Gemini 요약기 초기화
            gemini = None
            if use_gemini:
                try:
                    gemini = GeminiVideoSummarizer(gemini_key)
                    self.video_log_msg("✅ Gemini AI 연결 성공")
                except Exception as e:
                    self.video_log_msg(f"⚠️ Gemini 초기화 실패: {e}")
                    use_gemini = False
            
            all_videos = []
            all_comments = []
            
            for idx, (url, video_id) in enumerate(links, 1):
                self.video_log_msg(f"\n[{idx}/{len(links)}] 영상 분석 중...")
                self.video_log_msg(f"  URL: {url}")
                
                # 영상 정보 가져오기
                video_info = yt.get_video_info_by_id(video_id, self.video_log_msg)
                if not video_info:
                    self.video_log_msg(f"  ❌ 영상 정보를 가져올 수 없습니다.")
                    continue
                
                self.video_log_msg(f"  📺 제목: {video_info['영상 제목'][:40]}...")
                self.video_log_msg(f"  👁️ 조회수: {video_info['조회수']}")
                
                # Gemini AI 자막 기반 요약
                if use_gemini and gemini:
                    self.video_log_msg(f"  📝 자막 추출 중...")
                    transcript, lang, transcript_status = gemini.get_transcript(video_id)
                    
                    if transcript:
                        self.video_log_msg(f"  ✅ {transcript_status} 발견")
                        self.video_log_msg(f"  🤖 AI 요약 중...")
                        summary, status = gemini.summarize_with_transcript(
                            transcript, lang, video_info['영상 제목'], self.video_log_msg
                        )
                        if summary:
                            video_info['AI 요약'] = summary
                            video_info['요약 상태'] = f'성공 ({transcript_status})'
                            self.video_log_msg(f"  ✅ AI 요약 완료")
                        else:
                            video_info['AI 요약'] = ''
                            video_info['요약 상태'] = status
                            self.video_log_msg(f"  ⚠️ AI 요약: {status}")
                    else:
                        video_info['AI 요약'] = ''
                        video_info['요약 상태'] = transcript_status
                        self.video_log_msg(f"  ⚠️ {transcript_status}")
                    
                    time.sleep(2)  # Gemini API 할당량 보호 (분당 15회)
                else:
                    video_info['AI 요약'] = ''
                    video_info['요약 상태'] = 'AI 요약 미사용'
                
                all_videos.append(video_info)
                
                # 전체 댓글 수집
                if include_replies:
                    self.video_log_msg(f"  💬 댓글 + 대댓글 수집 중...")
                else:
                    self.video_log_msg(f"  💬 댓글 수집 중...")
                comments = yt.get_all_comments(video_id, video_info['영상 제목'], include_replies, self.video_log_msg)
                self.video_log_msg(f"  ✅ 댓글 {len(comments)}개 수집 완료")
                
                # 댓글 감성 분석 (Gemini 사용 시)
                if use_gemini and gemini and comments:
                    self.video_log_msg(f"  🎭 댓글 감성 분석 중...")
                    sentiment_result, sentiment_status = gemini.analyze_comments_sentiment(
                        comments, video_info['영상 제목'], self.video_log_msg
                    )
                    if sentiment_result:
                        video_info['댓글 감성 분석'] = sentiment_result
                        video_info['감성 분석 상태'] = '성공'
                        self.video_log_msg(f"  ✅ 댓글 감성 분석 완료")
                    else:
                        video_info['댓글 감성 분석'] = ''
                        video_info['감성 분석 상태'] = sentiment_status
                        self.video_log_msg(f"  ⚠️ 감성 분석: {sentiment_status}")
                    
                    time.sleep(2)  # Gemini API 할당량 보호
                else:
                    video_info['댓글 감성 분석'] = ''
                    video_info['감성 분석 상태'] = 'AI 분석 미사용' if not use_gemini else '댓글 없음'
                
                # 댓글에 video_id 추가하여 저장
                for c in comments:
                    c['video_id'] = video_id
                all_comments.extend(comments)
                
                # 진행률 업데이트
                progress = (idx / len(links)) * 100
                self.video_progress_var.set(progress)
                
                time.sleep(0.5)  # API 할당량 보호
            
            # 엑셀 저장 (세로 레포트 형식)
            self.video_log_msg("\n📊 엑셀 파일 생성 중 (세로 레포트 형식)...")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"유튜브영상분석_{timestamp}.xlsx"
            
            generator = VideoAnalysisExcelGenerator()
            generator.generate(all_videos, all_comments, filename)
            
            self.video_analysis_result = filename
            self.video_progress_var.set(100)
            
            self.video_log_msg("\n" + "=" * 50)
            self.video_log_msg(f"🎉 분석 완료!")
            self.video_log_msg(f"📊 영상: {len(all_videos)}개")
            self.video_log_msg(f"💬 댓글: {len(all_comments)}개")
            self.video_log_msg(f"📁 파일: {filename}")
            self.video_log_msg("=" * 50)
            
            self.root.after(0, lambda: self.video_start_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.video_open_btn.config(state=tk.NORMAL))
            
            messagebox.showinfo("완료", f"분석이 완료되었습니다!\n\n영상: {len(all_videos)}개\n댓글: {len(all_comments)}개\n\n파일: {filename}")
            
        except Exception as e:
            self.video_log_msg(f"\n❌ 오류 발생: {e}")
            self.root.after(0, lambda: self.video_start_btn.config(state=tk.NORMAL))
            messagebox.showerror("오류", f"분석 중 오류:\n{e}")
    
    def open_video_result(self):
        """영상 분석 결과 파일 열기"""
        if self.video_analysis_result and os.path.exists(self.video_analysis_result):
            os.startfile(self.video_analysis_result)
    
    def log(self, msg):
        self.log_text.insert(tk.END, f"{msg}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def live_log_msg(self, msg):
        self.live_log.insert(tk.END, f"{msg}\n")
        self.live_log.see(tk.END)
        self.root.update_idletasks()
    
    def start_analysis(self):
        keyword = self.keyword_var.get().strip()
        need_keyword = self.yt_check.get() or (self.dc_check.get() and self.dc_keyword_check.get())
        if need_keyword and not keyword:
            messagebox.showerror("오류", "유튜브 또는 디시 키워드 검색을 사용할 때는 분석 키워드를 입력해주세요!")
            return
        if not any([self.yt_check.get(), self.dc_check.get(), self.app_check.get(), self.play_check.get()]):
            messagebox.showerror("오류", "최소 1개 플랫폼을 선택해주세요!")
            return
        if self.dc_check.get():
            if not self.dc_keyword_check.get() and not self.dc_gallery_check.get():
                messagebox.showerror("오류", "디시인사이드: '키워드 검색 수집' 또는 '갤러리 글 수집' 중 하나 이상 선택해주세요!")
                return
            if self.dc_gallery_check.get() and not self.gallery_var.get().strip():
                messagebox.showerror("오류", "갤러리 글 수집을 선택했으면 갤러리 ID를 입력해주세요!")
                return
        if self.yt_check.get() and not self.api_key_var.get().strip():
            messagebox.showerror("오류", "YouTube API 키를 입력해주세요!")
            return
        if self.app_check.get() and not self.appstore_var.get().strip():
            messagebox.showerror("오류", "앱스토어 ID를 입력해주세요!")
            return
        if self.play_check.get() and not self.playstore_var.get().strip():
            messagebox.showerror("오류", "플레이스토어 ID를 입력해주세요!")
            return
        
        self.start_btn.config(state=tk.DISABLED)
        self.open_btn.config(state=tk.DISABLED)
        self.folder_btn.config(state=tk.DISABLED)
        self.log_text.delete(1.0, tk.END)
        self.progress_var.set(0)
        
        threading.Thread(target=self.run_analysis, daemon=True).start()
    
    def run_analysis(self):
        try:
            keyword = self.keyword_var.get().strip() or "동향분석"
            yt_count = int(self.yt_count_var.get())
            dc_pages = int(self.dc_page_var.get())
            
            self.log(f"🎯 '{keyword}' 분석 시작!")
            self.log("=" * 50)
            
            yt_videos, yt_comments, dc_posts, appstore_reviews, playstore_reviews = [], [], [], [], []
            total_steps = sum([self.yt_check.get(), self.dc_check.get(), self.app_check.get(), self.play_check.get()])
            step = 0
            
            if self.yt_check.get() and YOUTUBE_AVAILABLE:
                self.log(f"\n📺 유튜브 수집 시작 ({yt_count}개)...")
                try:
                    yt = YouTubeCrawler(self.api_key_var.get().strip())
                    yt_videos = yt.search_videos(keyword, yt_count, self.log)
                    self.log("  댓글 수집 중...")
                    for v in yt_videos[:10]:
                        yt_comments.extend(yt.get_comments(v['video_id'], v['영상 제목'], 50))
                        time.sleep(0.3)
                    self.log(f"✅ 유튜브: {len(yt_videos)}개 영상, {len(yt_comments)}개 댓글")
                except Exception as e:
                    self.log(f"❌ 유튜브 오류: {e}")
                step += 1
                self.progress_var.set(step / total_steps * 100)
            
            if self.dc_check.get():
                do_kw = self.dc_keyword_check.get()
                do_gal = self.dc_gallery_check.get()
                self.log(f"\n💬 디시인사이드 수집 시작 ({dc_pages}페이지)...")
                if do_kw:
                    self.log("  → 키워드 검색 수집 사용")
                if do_gal:
                    self.log("  → 갤러리 글 수집 사용")
                self.log("  ⚠️ IP 차단 방지를 위해 요청 간격을 두고 수집합니다.")
                try:
                    dc = DCInsideCrawler(days_limit=7)
                    gallery_id = self.gallery_var.get().strip() or None
                    dc_posts = dc.crawl_all(
                        keyword, gallery_id, self.is_minor_var.get(), dc_pages,
                        do_keyword=do_kw, do_gallery=do_gal, callback=self.log
                    )
                    self.log(f"✅ 디시인사이드: {len(dc_posts)}개 게시글")
                except Exception as e:
                    self.log(f"❌ 디시 오류: {e}")
                step += 1
                self.progress_var.set(step / total_steps * 100)
            
            if self.app_check.get():
                self.log("\n🍎 앱스토어 수집 시작...")
                try:
                    appstore_reviews = AppStoreCrawler().crawl(self.appstore_var.get().strip(), 10, self.log)
                    self.log(f"✅ 앱스토어: {len(appstore_reviews)}개 리뷰")
                except Exception as e:
                    self.log(f"❌ 앱스토어 오류: {e}")
                step += 1
                self.progress_var.set(step / total_steps * 100)
            
            if self.play_check.get() and PLAYSTORE_AVAILABLE:
                self.log("\n🤖 플레이스토어 수집 시작...")
                try:
                    playstore_reviews = PlayStoreCrawler().crawl(self.playstore_var.get().strip(), 500, self.log)
                    self.log(f"✅ 플레이스토어: {len(playstore_reviews)}개 리뷰")
                except Exception as e:
                    self.log(f"❌ 플레이스토어 오류: {e}")
                step += 1
                self.progress_var.set(step / total_steps * 100)
            
            self.log("\n📊 데이터 분석 중...")
            yt_summary = DataAnalyzer.summarize_youtube(yt_videos, yt_comments)
            dc_summary = DataAnalyzer.summarize_dc(dc_posts)
            app_summary = DataAnalyzer.summarize_reviews(appstore_reviews, '앱스토어')
            play_summary = DataAnalyzer.summarize_reviews(playstore_reviews, '플레이스토어')
            
            # AI 보고서 생성 (옵션)
            ai_report = None
            if self.ai_report_check.get():
                gemini_key = self.gemini_api_key_var.get().strip()
                if gemini_key and GEMINI_AVAILABLE:
                    self.log("\n🤖 AI 동향 분석 보고서 생성 중...")
                    try:
                        gemini = GeminiVideoSummarizer(gemini_key)
                        ai_report, status = gemini.generate_trend_report(
                            keyword, yt_summary, dc_summary, app_summary, play_summary, self.log
                        )
                        if ai_report:
                            self.log("✅ AI 보고서 생성 완료")
                        else:
                            self.log(f"⚠️ AI 보고서 생성 실패: {status}")
                    except Exception as e:
                        self.log(f"⚠️ AI 보고서 오류: {e}")
                else:
                    self.log("⚠️ AI 보고서: Gemini API 키가 없습니다. (유튜브 영상 분석 탭에서 설정)")
            
            self.log("📝 엑셀 파일 생성 중...")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            gen = ExcelGenerator(keyword)
            gen.timestamp = timestamp
            self.result_file = gen.generate(yt_videos, yt_comments, dc_posts, appstore_reviews, playstore_reviews,
                                            yt_summary, dc_summary, app_summary, play_summary, ai_report)
            
            if self.raw_check.get():
                self.log("\n📁 Raw 데이터 저장 중...")
                saver = RawDataSaver(keyword, timestamp)
                folder, files = saver.save_all(yt_videos, yt_comments, dc_posts, appstore_reviews, playstore_reviews)
                self.log(f"  → {len(files)}개 파일 저장됨")
            
            self.progress_var.set(100)
            self.log("\n" + "=" * 50)
            self.log("🎉 분석 완료!")
            self.log(f"📊 대시보드: {self.result_file}")
            self.log("=" * 50)
            
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.open_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.folder_btn.config(state=tk.NORMAL))
            messagebox.showinfo("완료", f"분석이 완료되었습니다!\n\n파일: {self.result_file}")
            
        except Exception as e:
            self.log(f"\n❌ 오류 발생: {e}")
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
            messagebox.showerror("오류", f"분석 중 오류:\n{e}")
    
    def start_live_collection(self):
        url = self.live_url_var.get().strip()
        if not url:
            messagebox.showerror("오류", "라이브 URL을 입력해주세요!")
            return
        
        platform = self.live_platform_var.get()
        
        if platform == "youtube" and not PYTCHAT_AVAILABLE:
            messagebox.showerror("오류", "pytchat 라이브러리가 설치되지 않았습니다.\npip install pytchat")
            return
        
        self.live_chats = []
        self.live_stop_event = threading.Event()
        self.live_log.delete(1.0, tk.END)
        
        self.live_start_btn.config(state=tk.DISABLED)
        self.live_stop_btn.config(state=tk.NORMAL)
        self.live_save_btn.config(state=tk.DISABLED)
        self.live_status_var.set("🔴 수집 중...")
        
        threading.Thread(target=self.run_live_collection, args=(platform, url), daemon=True).start()
    
    def run_live_collection(self, platform, url):
        try:
            if platform == "youtube":
                video_id = url.split("v=")[-1].split("&")[0] if "v=" in url else url
                crawler = YouTubeLiveCrawler()
                self.live_chats = crawler.collect(video_id, self.live_log_msg, self.live_stop_event)
            else:
                self.live_log_msg("❌ 지원하지 않는 플랫폼입니다.")
                return
        except Exception as e:
            self.live_log_msg(f"❌ 오류: {e}")
        finally:
            self.root.after(0, self.on_live_stopped)
    
    def stop_live_collection(self):
        if self.live_stop_event:
            self.live_stop_event.set()
        self.live_status_var.set("⏹ 중지됨")
    
    def on_live_stopped(self):
        self.live_start_btn.config(state=tk.NORMAL)
        self.live_stop_btn.config(state=tk.DISABLED)
        if self.live_chats:
            self.live_save_btn.config(state=tk.NORMAL)
        self.live_status_var.set(f"✅ 수집 완료 ({len(self.live_chats)}개)")
    
    def save_live_chats(self):
        if not self.live_chats:
            messagebox.showwarning("알림", "저장할 채팅이 없습니다.")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"live_chat_{timestamp}.csv"
        pd.DataFrame(self.live_chats).to_csv(filename, index=False, encoding='utf-8-sig')
        messagebox.showinfo("저장 완료", f"저장됨: {filename}\n총 {len(self.live_chats)}개 채팅")
    
    def open_result(self):
        if self.result_file and os.path.exists(self.result_file):
            os.startfile(self.result_file)
    
    def open_folder(self):
        os.startfile(os.getcwd())
    
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = AnalyzerApp()
    app.run()