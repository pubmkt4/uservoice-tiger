"""Excel 내보내기 — BytesIO로 생성하여 Streamlit 다운로드 버튼에 연결"""
import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── 컬러 팔레트 ───────────────────────────────────────────────
_C = {
    "header_bg":  "1565C0",
    "header_fg":  "FFFFFF",
    "pos_bg":     "C8E6C9",
    "neg_bg":     "FFCDD2",
    "neu_bg":     "F5F5F5",
    "summary_bg": "E3F2FD",
    "title_bg":   "0D47A1",
    "title_fg":   "FFFFFF",
}

_SENT_BG = {"긍정": _C["pos_bg"], "부정": _C["neg_bg"], "중립": _C["neu_bg"]}

_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(bold=True, color=_C["header_fg"])
        cell.fill      = _fill(_C["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value or ""))
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)


def _write_df(writer, df: pd.DataFrame, sheet_name: str,
              sentiment_col: str = None):
    """데이터프레임을 시트에 쓰고 스타일 적용"""
    if df.empty:
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    _style_header_row(ws, row=1, col_count=len(df.columns))

    # 감성 컬럼 컬러링
    if sentiment_col and sentiment_col in df.columns:
        sent_idx = df.columns.get_loc(sentiment_col) + 1
        for row_num in range(2, len(df) + 2):
            val  = ws.cell(row=row_num, column=sent_idx).value or ""
            bg   = _SENT_BG.get(val, _C["neu_bg"])
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=c)
                cell.fill   = _fill(bg)
                cell.border = _THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_summary(writer, keyword: str, insights: dict,
                   theme_stats: list, collection_result: dict):
    """요약 시트 작성"""
    ws = writer.book.create_sheet("📊 동향 요약", 0)

    # 타이틀
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = f"📊 {keyword} 동향 분석 리포트"
    title_cell.font      = Font(bold=True, size=16, color=_C["title_fg"])
    title_cell.fill      = _fill(_C["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws["A2"].value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font  = Font(italic=True, color="666666")

    # 수집 건수 요약
    row = 4
    ws.cell(row=row, column=1).value = "플랫폼"
    ws.cell(row=row, column=2).value = "수집 건수"
    _style_header_row(ws, row=row, col_count=2)
    row += 1
    counts = [
        ("YouTube 영상",    len(collection_result.get("yt_videos",   []))),
        ("YouTube 댓글",    len(collection_result.get("yt_comments", []))),
        ("YouTube 라이브채팅", len(collection_result.get("yt_live",  []))),
        ("디시인사이드",    len(collection_result.get("dc_posts",    []))),
        ("앱스토어",        len(collection_result.get("appstore",    []))),
        ("플레이스토어",    len(collection_result.get("playstore",   []))),
    ]
    for label, cnt in counts:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = cnt
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    # 종합 요약
    row += 1
    ws.cell(row=row, column=1).value = "종합 동향 요약"
    ws.cell(row=row, column=1).font  = Font(bold=True)
    row += 1
    summary_text = insights.get("summary", "분석 결과 없음")
    ws.merge_cells(f"A{row}:F{row+2}")
    cell = ws.cell(row=row, column=1)
    cell.value     = summary_text
    cell.fill      = _fill(_C["summary_bg"])
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    row += 4

    # 테마 통계
    if theme_stats:
        ws.cell(row=row, column=1).value = "주요 테마 분석"
        ws.cell(row=row, column=1).font  = Font(bold=True)
        row += 1
        headers = ["테마", "건수", "긍정%", "중립%", "부정%"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c).value = h
        _style_header_row(ws, row=row, col_count=len(headers))
        row += 1
        for stat in theme_stats:
            ws.cell(row=row, column=1).value = stat["name"]
            ws.cell(row=row, column=2).value = stat["count"]
            ws.cell(row=row, column=3).value = f"{stat['pos']}%"
            ws.cell(row=row, column=4).value = f"{stat['neu']}%"
            ws.cell(row=row, column=5).value = f"{stat['neg']}%"
            # 부정 비율 높으면 붉게
            if stat["neg"] >= 50:
                for c in range(1, 6):
                    ws.cell(row=row, column=c).fill = _fill(_C["neg_bg"])
            row += 1

    # 인사이트 카드
    row += 1
    ws.cell(row=row, column=1).value = "인사이트 카드"
    ws.cell(row=row, column=1).font  = Font(bold=True)
    row += 1
    card_headers = ["제목", "감성", "인사이트", "근거", "마케팅 시사점"]
    for c, h in enumerate(card_headers, 1):
        ws.cell(row=row, column=c).value = h
    _style_header_row(ws, row=row, col_count=len(card_headers))
    row += 1
    for card in insights.get("cards", []):
        ws.cell(row=row, column=1).value = card.get("title",   "")
        ws.cell(row=row, column=2).value = card.get("sentiment", "")
        ws.cell(row=row, column=3).value = card.get("insight",  "")
        ws.cell(row=row, column=4).value = card.get("evidence", "")
        ws.cell(row=row, column=5).value = card.get("action",   "")
        sent = card.get("sentiment", "혼재")
        bg = {"긍정": _C["pos_bg"], "부정": _C["neg_bg"]}.get(sent, _C["neu_bg"])
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill      = _fill(bg)
            cell.border    = _THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 50
        row += 1

    _auto_width(ws)


def generate_excel(keyword: str, collection_result: dict,
                   analysis_result: dict) -> bytes:
    """
    수집 + 분석 결과를 Excel 파일로 생성하여 bytes 반환.
    analysis_result가 없으면 raw 데이터만 포함.
    """
    from analysis.claude_analyzer import _aggregate_theme_stats

    insights      = (analysis_result or {}).get("insights", {})
    unified_themes = (analysis_result or {}).get("unified_themes", [])
    items_by_plat  = (analysis_result or {}).get("platform_items", {})
    theme_stats   = _aggregate_theme_stats(unified_themes, items_by_plat)

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # 데이터 시트 — 분석 결과가 있으면 sentiment/theme 컬럼 포함
        _write_data_sheets(writer, collection_result, items_by_plat)

        # 요약 시트는 마지막에 추가하고 나서 앞으로 이동
        _write_summary(writer, keyword, insights, theme_stats, collection_result)

    # 요약 시트를 맨 앞으로 이동
    wb = load_workbook(io.BytesIO(buffer.getvalue()))
    sheets = wb.sheetnames
    if "📊 동향 요약" in sheets:
        wb.move_sheet("📊 동향 요약", offset=-len(sheets))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_data_sheets(writer, collection_result: dict, items_by_plat: dict):
    """플랫폼별 데이터 시트 작성"""

    def _merge_analysis(raw: list, analyzed: list, id_keys: list) -> list:
        """분석 결과(sentiment, theme)를 raw 데이터에 병합"""
        if not analyzed:
            return raw
        # ID 기반 인덱스 구성
        idx = {}
        for item in analyzed:
            key = tuple(str(item.get(k, "")) for k in id_keys)
            idx[key] = item
        merged = []
        for item in raw:
            key  = tuple(str(item.get(k, "")) for k in id_keys)
            base = dict(item)
            if key in idx:
                base["sentiment"] = idx[key].get("sentiment", "")
                base["theme"]     = idx[key].get("theme", "")
            merged.append(base)
        return merged

    # YouTube 영상
    yt_videos = collection_result.get("yt_videos", [])
    if yt_videos:
        cols = ["순위", "영상 제목", "채널명", "채널 구독자",
                "조회수", "좋아요", "댓글수", "업로드일", "링크"]
        df = pd.DataFrame(yt_videos)[[c for c in cols if c in pd.DataFrame(yt_videos).columns]]
        _write_df(writer, df, "YouTube_영상")

    # YouTube 댓글
    yt_comments = collection_result.get("yt_comments", [])
    analyzed_cmt = items_by_plat.get("YouTube 댓글", [])
    if yt_comments:
        merged = _merge_analysis(yt_comments, analyzed_cmt, ["video_id", "작성자", "작성일"])
        cols = ["영상 제목", "작성자", "댓글", "좋아요", "작성일", "유형", "sentiment", "theme"]
        df = pd.DataFrame(merged)[[c for c in cols if c in pd.DataFrame(merged).columns]]
        df.rename(columns={"sentiment": "감성", "theme": "테마"}, inplace=True)
        _write_df(writer, df, "YouTube_댓글", sentiment_col="감성")

    # 디시인사이드
    dc_posts = collection_result.get("dc_posts", [])
    analyzed_dc = items_by_plat.get("디시인사이드", [])
    if dc_posts:
        merged = _merge_analysis(dc_posts, analyzed_dc, ["링크"])
        cols = ["제목", "본문", "작성일", "출처", "링크", "sentiment", "theme"]
        df = pd.DataFrame(merged)[[c for c in cols if c in pd.DataFrame(merged).columns]]
        df.rename(columns={"sentiment": "감성", "theme": "테마"}, inplace=True)
        _write_df(writer, df, "디시인사이드", sentiment_col="감성")

    # 앱스토어
    appstore = collection_result.get("appstore", [])
    analyzed_app = items_by_plat.get("앱스토어", [])
    if appstore:
        merged = _merge_analysis(appstore, analyzed_app, ["review_id"])
        cols = ["제목", "작성자", "평점", "내용", "작성일", "버전", "sentiment", "theme"]
        df = pd.DataFrame(merged)[[c for c in cols if c in pd.DataFrame(merged).columns]]
        df.rename(columns={"sentiment": "감성", "theme": "테마"}, inplace=True)
        _write_df(writer, df, "앱스토어_리뷰", sentiment_col="감성")

    # 플레이스토어
    playstore = collection_result.get("playstore", [])
    analyzed_play = items_by_plat.get("플레이스토어", [])
    if playstore:
        merged = _merge_analysis(playstore, analyzed_play, ["작성자", "작성일"])
        cols = ["작성자", "평점", "내용", "작성일", "좋아요", "sentiment", "theme"]
        df = pd.DataFrame(merged)[[c for c in cols if c in pd.DataFrame(merged).columns]]
        df.rename(columns={"sentiment": "감성", "theme": "테마"}, inplace=True)
        _write_df(writer, df, "플레이스토어_리뷰", sentiment_col="감성")

    # YouTube 라이브채팅
    yt_live = collection_result.get("yt_live", [])
    analyzed_live = items_by_plat.get("YouTube 라이브채팅", [])
    if yt_live:
        merged = _merge_analysis(yt_live, analyzed_live, ["작성자", "시간"])
        cols = ["시간", "작성자", "메시지", "sentiment", "theme"]
        df = pd.DataFrame(merged)[[c for c in cols if c in pd.DataFrame(merged).columns]]
        df.rename(columns={"sentiment": "감성", "theme": "테마"}, inplace=True)
        _write_df(writer, df, "YouTube_라이브채팅", sentiment_col="감성")
